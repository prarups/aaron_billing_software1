from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
from core.models import Product, Branch, ProductRegistry, StockTransaction, ComboGroup

User = get_user_model()

class StockPivotReportTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.branch = Branch.objects.create(name="Nellore Branch", code="10001", invoice_prefix="AN")
        self.user = User.objects.create_user(
            username="testowner", 
            password="password123", 
            role="owner",
            active_branch=self.branch
        )
        self.user.branches.add(self.branch)
        self.user.save()
        self.client.login(username="testowner", password="password123")

        self.product = Product.objects.create(name="jeans", barcode="610016", price=500, branch=self.branch)
        self.registry = ProductRegistry.objects.create(
            branch=self.branch,
            product=self.product,
            stock_quantity=88
        )

        # Create some historical/all-time stock transactions
        # Out (sales): 9
        StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='OUT',
            quantity=9,
            user=self.user
        )
        # Dmg: 2
        StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='DMG',
            quantity=2,
            user=self.user
        )
        # So current stock is 88. All-time OUT is 9. All-time DMG is 2.
        # Total received all-time should be: 88 + 9 + 2 = 99.

    def test_stock_pivot_report_view(self):
        # Create an ADJ transaction today of -5 items and reduce registry stock accordingly
        self.registry.stock_quantity = 83
        self.registry.save()
        
        StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='ADJ',
            quantity=-5,
            user=self.user
        )

        response = self.client.get(reverse('stock_pivot_report'))
        self.assertEqual(response.status_code, 200)
        
        report_data = response.context['report_data']
        self.assertEqual(len(report_data), 1)
        item = report_data[0]
        self.assertEqual(item['product'], self.product)
        self.assertEqual(item['total_rec'], 94)
        self.assertEqual(item['total_all_time_out'], 9)
        self.assertEqual(item['total_all_time_dmg'], 2)
        self.assertEqual(item['total_adj_plus'], 0)
        self.assertEqual(item['total_adj_minus'], 5)

        branch_stock = item['branch_stocks'][0]
        self.assertEqual(branch_stock['rec'], 94)
        self.assertEqual(branch_stock['all_time_out'], 9)
        self.assertEqual(branch_stock['all_time_dmg'], 2)
        self.assertEqual(branch_stock['adj_plus'], 0)
        self.assertEqual(branch_stock['adj_minus'], 5)

    def test_stock_pivot_report_view_with_returns(self):
        # Create a GOOD return of 3 items
        StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='IN',
            quantity=3,
            reference='Return #123 (GOOD)',
            user=self.user
        )
        # Create a DAMAGED return of 1 item
        StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='DMG',
            quantity=1,
            reference='Return #123 (DAMAGED)',
            user=self.user
        )
        
        # Initial stock: 88. Let's set it to 85.
        self.registry.stock_quantity = 85
        self.registry.save()

        response = self.client.get(reverse('stock_pivot_report'))
        self.assertEqual(response.status_code, 200)

        report_data = response.context['report_data']
        self.assertEqual(len(report_data), 1)
        item = report_data[0]
        
        # Let's verify formulas:
        # curr_stock = 85
        # all_time_out_gross = 9 (from setUp)
        # all_time_dmg_std = 2 (from setUp)
        # all_time_ret = 3 (GOOD return)
        # all_time_ret_dmg = 1 (DAMAGED return)
        # all_time_adj = 0
        #
        # total_rec = curr_stock + all_time_out_gross + all_time_dmg_std - all_time_ret - all_time_adj
        #           = 85 + 9 + 2 - 3 - 0 = 93.
        #
        # total_all_time_out = all_time_out_gross - all_time_ret - all_time_ret_dmg
        #                    = 9 - 3 - 1 = 5.
        
        self.assertEqual(item['total_rec'], 93)
        self.assertEqual(item['total_all_time_out'], 5)
        self.assertEqual(item['total_all_time_dmg'], 3)
        self.assertEqual(item['total_ret'], 3)
        self.assertEqual(item['total_ret_dmg'], 1)

        branch_stock = item['branch_stocks'][0]
        self.assertEqual(branch_stock['rec'], 93)
        self.assertEqual(branch_stock['all_time_out'], 5)
        self.assertEqual(branch_stock['all_time_dmg'], 3)
        self.assertEqual(branch_stock['ret'], 3)
        self.assertEqual(branch_stock['ret_dmg'], 1)

    def test_export_stock_pivot_excel(self):
        response = self.client.get(reverse('export_stock_pivot_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_product_update_stock_add(self):
        # Initial stock was 88
        url = reverse('product_update', args=[self.product.pk]) + f"?reg_id={self.registry.pk}"
        post_data = {
            'name': 'jeans',
            'barcode': '610016',
            'price': '500',
            'low_stock_threshold': '10',
            'stock_update_type': 'add',
            'stock_update_qty': '12',
            'stock_update_reason': 'Arrived today',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        
        self.registry.refresh_from_db()
        self.assertEqual(self.registry.stock_quantity, 100) # 88 + 12

        # Check transaction
        tx = StockTransaction.objects.filter(transaction_type='IN', reference__contains='Arrived today').first()
        self.assertIsNotNone(tx)
        self.assertEqual(tx.quantity, 12)

    def test_product_update_stock_damage(self):
        # Initial stock was 88
        url = reverse('product_update', args=[self.product.pk]) + f"?reg_id={self.registry.pk}"
        post_data = {
            'name': 'jeans',
            'barcode': '610016',
            'price': '500',
            'low_stock_threshold': '10',
            'stock_update_type': 'damage',
            'stock_update_qty': '4',
            'stock_update_reason': 'Moth eaten',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        
        self.registry.refresh_from_db()
        self.assertEqual(self.registry.stock_quantity, 84) # 88 - 4

        # Check transaction
        tx = StockTransaction.objects.filter(transaction_type='DMG', reference__contains='Moth eaten').first()
        self.assertIsNotNone(tx)
        self.assertEqual(tx.quantity, 4)

    def test_product_update_stock_correction(self):
        # Clear existing transactions to prevent interference
        StockTransaction.objects.all().delete()
        
        # Initial stock was 88
        url = reverse('product_update', args=[self.product.pk]) + f"?reg_id={self.registry.pk}"
        post_data = {
            'name': 'jeans',
            'barcode': '610016',
            'price': '500',
            'low_stock_threshold': '10',
            'stock_update_type': 'correction',
            'stock_update_qty': '95',
            'stock_update_reason': 'Audit match',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        
        self.registry.refresh_from_db()
        self.assertEqual(self.registry.stock_quantity, 95)

        # Check transaction
        tx = StockTransaction.objects.filter(transaction_type='IN', reference__contains='Audit match').first()
        self.assertIsNotNone(tx)
        self.assertEqual(tx.quantity, 7) # 95 - 88 = 7

    def test_product_update_stock_negative_correction(self):
        # Clear existing transactions to prevent interference
        StockTransaction.objects.all().delete()

        # Initial stock was 88
        url = reverse('product_update', args=[self.product.pk]) + f"?reg_id={self.registry.pk}"
        post_data = {
            'name': 'jeans',
            'barcode': '610016',
            'price': '500',
            'low_stock_threshold': '10',
            'stock_update_type': 'correction',
            'stock_update_qty': '80',
            'stock_update_reason': 'Audit match negative',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        
        self.registry.refresh_from_db()
        self.assertEqual(self.registry.stock_quantity, 80)

        # Check transaction
        tx = StockTransaction.objects.filter(transaction_type='ADJ', reference__contains='Audit match negative').first()
        self.assertIsNotNone(tx)
        self.assertEqual(tx.quantity, -8) # 80 - 88 = -8

    def test_product_update_stock_correction_with_today_in_txn(self):
        # Clear existing transactions to prevent interference
        StockTransaction.objects.all().delete()

        # Create an IN transaction today of 10 items
        StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='IN',
            quantity=10,
            user=self.user
        )
        # Initial stock was 88.
        # Now correct it down to 85 (diff is -3).
        url = reverse('product_update', args=[self.product.pk]) + f"?reg_id={self.registry.pk}"
        post_data = {
            'name': 'jeans',
            'barcode': '610016',
            'price': '500',
            'low_stock_threshold': '10',
            'stock_update_type': 'correction',
            'stock_update_qty': '85',
            'stock_update_reason': 'Correct receipt typo',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        
        self.registry.refresh_from_db()
        self.assertEqual(self.registry.stock_quantity, 85)

        # The IN transaction today should have been reduced from 10 to 7!
        tx = StockTransaction.objects.filter(transaction_type='IN').first()
        self.assertIsNotNone(tx)
        self.assertEqual(tx.quantity, 7)

        # No ADJ transaction should be created because it was absorbed by the IN transaction
        adj_tx = StockTransaction.objects.filter(transaction_type='ADJ').first()
        self.assertIsNone(adj_tx)

    def test_product_update_stock_correction_of_wrong_damage(self):
        # Clear existing transactions to prevent interference
        StockTransaction.objects.all().delete()

        # Create a DMG transaction today of 5 items
        # Let's say stock went from 88 to 83.
        self.registry.stock_quantity = 83
        self.registry.save()
        
        StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='DMG',
            quantity=5,
            user=self.user
        )

        # Now correct stock back to 85 (diff is +2).
        url = reverse('product_update', args=[self.product.pk]) + f"?reg_id={self.registry.pk}"
        post_data = {
            'name': 'jeans',
            'barcode': '610016',
            'price': '500',
            'low_stock_threshold': '10',
            'stock_update_type': 'correction',
            'stock_update_qty': '85',
            'stock_update_reason': 'Correct wrong damage entry',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        
        self.registry.refresh_from_db()
        self.assertEqual(self.registry.stock_quantity, 85)

        # Today's DMG transaction should have been reduced from 5 to 3!
        tx = StockTransaction.objects.filter(transaction_type='DMG', reference__contains='Correct wrong damage entry').first()
        self.assertIsNotNone(tx)
        self.assertEqual(tx.quantity, 3)

        # No IN or ADJ transaction should be created because it was absorbed by the DMG correction
        in_tx = StockTransaction.objects.filter(transaction_type='IN', reference__contains='Correct wrong damage entry').first()
        self.assertIsNone(in_tx)
        adj_tx = StockTransaction.objects.filter(transaction_type='ADJ', reference__contains='Correct wrong damage entry').first()
        self.assertIsNone(adj_tx)

    def test_product_update_stock_correct_damage_increase(self):
        # Clear existing transactions to prevent interference
        StockTransaction.objects.all().delete()

        # Initial stock was 88. Let's add 2 DMG transactions previously.
        StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='DMG',
            quantity=2,
            user=self.user
        )
        self.registry.stock_quantity = 88
        self.registry.damaged_qty = 2
        self.registry.save()

        # We set new damaged total to 5 (increase damage by 3).
        # Sellable stock should decrease by 3 (from 88 to 85).
        url = reverse('product_update', args=[self.product.pk]) + f"?reg_id={self.registry.pk}"
        post_data = {
            'name': 'jeans',
            'barcode': '610016',
            'price': '500',
            'low_stock_threshold': '10',
            'stock_update_type': 'correct_damage',
            'stock_update_qty': '5',
            'stock_update_reason': 'More damages found',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)

        self.registry.refresh_from_db()
        self.assertEqual(self.registry.stock_quantity, 85)
        self.assertEqual(self.registry.damaged_qty, 5)

        # There should be DMG transaction with quantity = 3 (or today's dmg transaction modified to sum to 5)
        from django.db.models import Sum
        dmg_sum = StockTransaction.objects.filter(product=self.product, branch=self.branch, transaction_type='DMG').aggregate(t=Sum('quantity'))['t']
        self.assertEqual(dmg_sum, 5)

    def test_product_update_stock_correct_damage_decrease_with_today_dmg_txn(self):
        # Clear existing transactions to prevent interference
        StockTransaction.objects.all().delete()

        # Create DMG transaction today of 5 items
        self.registry.stock_quantity = 83
        self.registry.damaged_qty = 5
        self.registry.save()
        StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='DMG',
            quantity=5,
            user=self.user
        )

        # Set new damaged total to 3 (decrease damage by 2).
        # Sellable stock should increase by 2 (from 83 to 85).
        url = reverse('product_update', args=[self.product.pk]) + f"?reg_id={self.registry.pk}"
        post_data = {
            'name': 'jeans',
            'barcode': '610016',
            'price': '500',
            'low_stock_threshold': '10',
            'stock_update_type': 'correct_damage',
            'stock_update_qty': '3',
            'stock_update_reason': 'Fewer damages actual',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)

        self.registry.refresh_from_db()
        self.assertEqual(self.registry.stock_quantity, 85)
        self.assertEqual(self.registry.damaged_qty, 3)

        # Today's DMG transaction should have been reduced to 3
        tx = StockTransaction.objects.filter(transaction_type='DMG').first()
        self.assertEqual(tx.quantity, 3)

    def test_product_update_stock_correct_damage_decrease_no_today_dmg_txn(self):
        # Clear existing transactions to prevent interference
        StockTransaction.objects.all().delete()

        # Setup historic DMG transactions (e.g. from yesterday)
        import datetime
        from django.utils import timezone
        yesterday = timezone.now() - datetime.timedelta(days=1)
        
        tx = StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='DMG',
            quantity=5,
            user=self.user
        )
        # Manually force created_at to yesterday
        StockTransaction.objects.filter(pk=tx.pk).update(created_at=yesterday)

        self.registry.stock_quantity = 83
        self.registry.damaged_qty = 5
        self.registry.save()

        # Set new damaged total to 3 (decrease damage by 2).
        # Sellable stock should increase by 2 (from 83 to 85).
        url = reverse('product_update', args=[self.product.pk]) + f"?reg_id={self.registry.pk}"
        post_data = {
            'name': 'jeans',
            'barcode': '610016',
            'price': '500',
            'low_stock_threshold': '10',
            'stock_update_type': 'correct_damage',
            'stock_update_qty': '3',
            'stock_update_reason': 'Historic damage correction',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)

        self.registry.refresh_from_db()
        self.assertEqual(self.registry.stock_quantity, 85)
        self.assertEqual(self.registry.damaged_qty, 3)

        # A new DMG transaction with quantity -2 should have been logged
        new_tx = StockTransaction.objects.filter(transaction_type='DMG', quantity=-2).first()
        self.assertIsNotNone(new_tx)

    def test_product_update_stock_correct_damage_negative_stock_error(self):
        # Clear existing transactions to prevent interference
        StockTransaction.objects.all().delete()

        # Setup 2 sellable, 1 damaged
        self.registry.stock_quantity = 2
        self.registry.damaged_qty = 1
        self.registry.save()
        StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='DMG',
            quantity=1,
            user=self.user
        )

        # Correct total damage to 5 (needs 4 more units, but only 2 sellable exist).
        # This should fail with validation error and not redirect.
        url = reverse('product_update', args=[self.product.pk]) + f"?reg_id={self.registry.pk}"
        post_data = {
            'name': 'jeans',
            'barcode': '610016',
            'price': '500',
            'low_stock_threshold': '10',
            'stock_update_type': 'correct_damage',
            'stock_update_qty': '5',
            'stock_update_reason': 'Invalid damage count',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200) # Form page returned with errors

        self.registry.refresh_from_db()
        self.assertEqual(self.registry.stock_quantity, 2)
        self.assertEqual(self.registry.damaged_qty, 1)

    def test_stock_transaction_and_adjustment_truncation(self):
        # Create a transaction with a very long reference (> 100 characters)
        long_ref = "A" * 150
        tx = StockTransaction.objects.create(
            product=self.product,
            branch=self.branch,
            transaction_type='IN',
            quantity=10,
            reference=long_ref,
            user=self.user
        )
        self.assertEqual(len(tx.reference), 100)
        self.assertEqual(tx.reference, "A" * 100)

        # Create an adjustment with a very long reason (> 255 characters)
        from core.models import StockAdjustment
        long_reason = "B" * 300
        adj = StockAdjustment.objects.create(
            product=self.product,
            branch=self.branch,
            opening_balance=10,
            stock_in=0,
            stock_out=0,
            correction_amount=5,
            closing_stock=15,
            is_in_stock=True,
            reason=long_reason,
            user=self.user
        )
        self.assertEqual(len(adj.reason), 255)
        self.assertEqual(adj.reason, "B" * 255)

    def test_stock_pivot_report_view_excludes_adjustment_columns(self):
        # Verify that adjustment columns are removed visually from HTML,
        # but the View Corrected Details link exists for the branch breakdown.
        response = self.client.get(reverse('stock_pivot_report') + f"?branch={self.branch.id}")
        self.assertEqual(response.status_code, 200)
        html = response.content.decode('utf-8')
        
        # Verify columns '+ Adj' and '- Adj' are not in headers
        self.assertNotIn('+ Adj', html)
        self.assertNotIn('- Adj', html)

        # Verify the history button links to our new route
        history_url = reverse('view_stock_adjustments', args=[self.registry.pk])
        self.assertIn(history_url, html)
        self.assertIn('bi-clock-history', html)

    def test_view_stock_adjustments_history_view_access(self):
        # Create an adjustment
        from core.models import StockAdjustment
        adj = StockAdjustment.objects.create(
            product=self.product,
            branch=self.branch,
            opening_balance=88,
            stock_in=0,
            stock_out=0,
            correction_amount=10,
            closing_stock=98,
            is_in_stock=True,
            reason="Test manual fix",
            user=self.user
        )

        history_url = reverse('view_stock_adjustments', args=[self.registry.pk])
        
        # 1. Owner/Manager (self.user) gets access and sees the adjustment
        response = self.client.get(history_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Test manual fix")
        self.assertContains(response, "+10")
        
        # 2. Staff user is blocked (redirected)
        staff_user = User.objects.create_user(
            username="teststaff", 
            password="password123", 
            role="staff",
            active_branch=self.branch
        )
        self.client.force_login(staff_user)
        response = self.client.get(history_url)
        self.assertEqual(response.status_code, 302) # Redirect to dashboard
        
        # 3. User without branch access is blocked
        other_branch = Branch.objects.create(name="Tirupati Branch", code="10002", invoice_prefix="TP")
        other_manager = User.objects.create_user(
            username="othermanager", 
            password="password123", 
            role="manager",
            active_branch=other_branch
        )
        other_manager.branches.add(other_branch)
        other_manager.save()
        
        self.client.force_login(other_manager)
        response = self.client.get(history_url)
        self.assertRedirects(response, reverse('stock_pivot_report'))

    def test_stock_pivot_report_view_with_branch_filter(self):
        # Create a second branch
        branch2 = Branch.objects.create(name="Nellore Branch 2", code="10002", invoice_prefix="AN2")
        # Register the product to branch2
        ProductRegistry.objects.create(
            branch=branch2,
            product=self.product,
            stock_quantity=45
        )

        # 1. Fetching without filter should return both branches' stocks
        response = self.client.get(reverse('stock_pivot_report'))
        self.assertEqual(response.status_code, 200)
        report_data = response.context['report_data']
        self.assertEqual(len(report_data), 1)
        self.assertEqual(len(report_data[0]['branch_stocks']), 2)

        # 2. Filter by first branch (self.branch)
        response_f1 = self.client.get(reverse('stock_pivot_report') + f"?branch={self.branch.id}")
        self.assertEqual(response_f1.status_code, 200)
        report_data_f1 = response_f1.context['report_data']
        self.assertEqual(len(report_data_f1), 1)
        self.assertEqual(len(report_data_f1[0]['branch_stocks']), 1)
        self.assertEqual(report_data_f1[0]['branch_stocks'][0]['branch'], self.branch)
        self.assertEqual(report_data_f1[0]['branch_stocks'][0]['cl'], 88)

        # 3. Filter by second branch (branch2)
        response_f2 = self.client.get(reverse('stock_pivot_report') + f"?branch={branch2.id}")
        self.assertEqual(response_f2.status_code, 200)
        report_data_f2 = response_f2.context['report_data']
        self.assertEqual(len(report_data_f2), 1)
        self.assertEqual(len(report_data_f2[0]['branch_stocks']), 1)
        self.assertEqual(report_data_f2[0]['branch_stocks'][0]['branch'], branch2)
        self.assertEqual(report_data_f2[0]['branch_stocks'][0]['cl'], 45)

    def test_export_stock_pivot_excel_with_branch_filter(self):
        # Create a second branch
        branch2 = Branch.objects.create(name="Nellore Branch 2", code="10002", invoice_prefix="AN2")
        ProductRegistry.objects.create(
            branch=branch2,
            product=self.product,
            stock_quantity=45
        )

        # Export Excel filtered by branch2
        response = self.client.get(reverse('export_stock_pivot_excel') + f"?branch={branch2.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Read the excel workbook to verify it only exported details for branch2
        import openpyxl
        import io
        wb = openpyxl.load_workbook(filename=io.BytesIO(response.content))
        
        # In Sheet 2 (Branch Details), row 2 should contain "Nellore Branch 2"
        ws2 = wb['Branch Details']
        rows = list(ws2.iter_rows(values_only=True))
        
        # Header is row 0, data starts at row 1.
        self.assertEqual(len(rows), 2) # Header + 1 data row
        self.assertEqual(rows[1][2], "Nellore Branch 2") # Branch Name column is index 2

    def test_stock_pivot_report_displays_branch_product_count(self):
        # The user has access to one branch (self.branch) with 1 product registered
        response = self.client.get(reverse('stock_pivot_report'))
        self.assertEqual(response.status_code, 200)
        
        # Verify the context contains accessible_branches with annotated product_count
        accessible_branches = response.context['accessible_branches']
        self.assertTrue(any(b.id == self.branch.id and b.product_count == 1 for b in accessible_branches))
        
        # Verify the HTML contains the correct label and value
        self.assertContains(response, "Products: 1")


class BulkInsertTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.branch = Branch.objects.create(name="Nellore Branch", code="10001", invoice_prefix="AN")
        self.user = User.objects.create_user(
            username="testowner", 
            password="password123", 
            role="owner",
            active_branch=self.branch
        )
        self.user.branches.add(self.branch)
        self.user.save()
        self.client.login(username="testowner", password="password123")

    def test_bulk_insert_success(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_content = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            "T-Shirt,TSHIRT01,250.00,M,10001,15,5\n"
            "Jeans,JEANS01,500.00,L,10001,20,10\n"
        )
        csv_file = SimpleUploadedFile("bulk.csv", csv_content.encode("utf-8"), content_type="text/csv")
        response = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file})
        self.assertEqual(response.status_code, 200)

        # Check product registry is created
        product = Product.objects.get(barcode="TSHIRT01")
        reg = ProductRegistry.objects.get(product=product, branch=self.branch)
        self.assertEqual(reg.stock_quantity, 15)

        product2 = Product.objects.get(barcode="JEANS01")
        reg2 = ProductRegistry.objects.get(product=product2, branch=self.branch)
        self.assertEqual(reg2.stock_quantity, 20)

    def test_bulk_insert_missing_mandatory_fields(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        # Missing Barcode (mandatory)
        csv_content1 = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            "T-Shirt,,250.00,M,10001,15,5\n"
        )
        csv_file1 = SimpleUploadedFile("bulk1.csv", csv_content1.encode("utf-8"), content_type="text/csv")
        response1 = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file1})
        self.assertEqual(response1.status_code, 200)
        self.assertFalse(Product.objects.filter(name="T-Shirt").exists())

        # Missing Initial Stock (mandatory)
        csv_content2 = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            "T-Shirt,TSHIRT01,250.00,M,10001,,5\n"
        )
        csv_file2 = SimpleUploadedFile("bulk2.csv", csv_content2.encode("utf-8"), content_type="text/csv")
        response2 = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file2})
        self.assertEqual(response2.status_code, 200)
        self.assertFalse(Product.objects.filter(barcode="TSHIRT01").exists())

    def test_bulk_insert_optional_name_and_size(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        # Name and Size are optional
        csv_content = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            ",TSHIRT_NO_NAME,250.00,,10001,15,5\n"
        )
        csv_file = SimpleUploadedFile("bulk_opt.csv", csv_content.encode("utf-8"), content_type="text/csv")
        response = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file})
        self.assertEqual(response.status_code, 200)
        
        # Product should be successfully created with empty name and size
        self.assertTrue(Product.objects.filter(barcode="TSHIRT_NO_NAME").exists())
        product = Product.objects.get(barcode="TSHIRT_NO_NAME")
        self.assertEqual(product.name, "")
        self.assertEqual(product.size, "")

    def test_bulk_insert_duplicate_barcode_in_file(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_content = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            "T-Shirt,TSHIRT01,250.00,M,10001,15,5\n"
            "Another T-Shirt,TSHIRT01,300.00,L,10001,10,5\n"
        )
        csv_file = SimpleUploadedFile("bulk.csv", csv_content.encode("utf-8"), content_type="text/csv")
        response = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file})
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Product.objects.filter(barcode="TSHIRT01").exists())

    def test_bulk_insert_duplicate_name_in_file(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_content = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            "T-Shirt,TSHIRT01,250.00,M,10001,15,5\n"
            "T-Shirt,TSHIRT02,300.00,L,10001,10,5\n"
        )
        csv_file = SimpleUploadedFile("bulk.csv", csv_content.encode("utf-8"), content_type="text/csv")
        response = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(Product.objects.filter(barcode="TSHIRT01").exists())
        self.assertTrue(Product.objects.filter(barcode="TSHIRT02").exists())

    def test_bulk_insert_duplicate_barcode_in_database(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        # Pre-create a product and register it for the branch
        product = Product.objects.create(name="T-Shirt", barcode="TSHIRT01", price=250, branch=self.branch)
        ProductRegistry.objects.create(branch=self.branch, product=product, stock_quantity=10)
        
        csv_content = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            "T-Shirt,TSHIRT01,250.00,M,10001,15,5\n"
        )
        csv_file = SimpleUploadedFile("bulk.csv", csv_content.encode("utf-8"), content_type="text/csv")
        response = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file})
        self.assertEqual(response.status_code, 200)
        
        # Verify it wasn't modified or stock added, and the error was raised
        product.refresh_from_db()
        reg = ProductRegistry.objects.get(product=product, branch=self.branch)
        self.assertEqual(reg.stock_quantity, 10) # Unchanged

    def test_bulk_insert_duplicate_name_in_database(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        # Pre-create a product with same name but different barcode and register it for the branch
        product = Product.objects.create(name="T-Shirt", barcode="TSHIRT02", price=250, branch=self.branch)
        ProductRegistry.objects.create(branch=self.branch, product=product, stock_quantity=10)
        
        csv_content = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            "T-Shirt,TSHIRT01,250.00,M,10001,15,5\n"
        )
        csv_file = SimpleUploadedFile("bulk.csv", csv_content.encode("utf-8"), content_type="text/csv")
        response = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(Product.objects.filter(barcode="TSHIRT01").exists())

    def test_bulk_insert_same_barcode_different_branch_allowed(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        # Create second branch
        other_branch = Branch.objects.create(name="Tirupati Branch", code="10002", invoice_prefix="TP")
        self.user.branches.add(other_branch)
        
        # Pre-create product and register for first branch (Nellore)
        product = Product.objects.create(name="T-Shirt", barcode="TSHIRT01", price=250, branch=self.branch)
        ProductRegistry.objects.create(branch=self.branch, product=product, stock_quantity=10)
        
        # Uploading same barcode and name for the second branch (Tirupati) should succeed
        csv_content = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            "T-Shirt,TSHIRT01,250.00,M,10002,15,5\n"
        )
        csv_file = SimpleUploadedFile("bulk.csv", csv_content.encode("utf-8"), content_type="text/csv")
        response = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file})
        self.assertEqual(response.status_code, 200)
        
        # Verify it created a new product and registry for the other branch
        new_prod = Product.objects.get(barcode="TSHIRT01", branch=other_branch)
        reg_other = ProductRegistry.objects.get(product=new_prod, branch=other_branch)
        self.assertEqual(reg_other.stock_quantity, 15)

    def test_bulk_insert_missing_price(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_content = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            "T-Shirt,TSHIRT01,,M,10001,15,5\n"
        )
        csv_file = SimpleUploadedFile("bulk.csv", csv_content.encode("utf-8"), content_type="text/csv")
        response = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file})
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Product.objects.filter(barcode="TSHIRT01").exists())

    def test_bulk_insert_negative_price(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_content = (
            "Name,Barcode,Price,Size,Branch Code,Initial Stock,Low Stock Alert\n"
            "T-Shirt,TSHIRT01,-15.50,M,10001,15,5\n"
        )
        csv_file = SimpleUploadedFile("bulk.csv", csv_content.encode("utf-8"), content_type="text/csv")
        response = self.client.post(reverse('bulk_insert'), {'csv_file': csv_file})
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Product.objects.filter(barcode="TSHIRT01").exists())

    def test_bulk_insert_excel_success(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        import openpyxl
        import io
        
        # Create an in-memory workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Barcode", "Price", "Size", "Branch Code", "Initial Stock", "Low Stock Alert"])
        # Use floating point numbers in Excel (e.g. 10001.0, 25.0) to test safety of conversions
        ws.append(["T-Shirt Excel", "EXCELTSHIRT01", 350.00, "M", 10001, 25, 5])
        ws.append(["Jeans Excel", "EXCELJEANS01", 600.00, "L", 10001.0, 30.0, 10.0])
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        
        excel_file = SimpleUploadedFile(
            "bulk.xlsx",
            excel_io.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(reverse('bulk_insert'), {'csv_file': excel_file})
        self.assertEqual(response.status_code, 200)
        
        # Check product registry is created
        product = Product.objects.get(barcode="EXCELTSHIRT01")
        reg = ProductRegistry.objects.get(product=product, branch=self.branch)
        self.assertEqual(reg.stock_quantity, 25)
        self.assertEqual(product.price, 350.00)
        
        product2 = Product.objects.get(barcode="EXCELJEANS01")
        reg2 = ProductRegistry.objects.get(product=product2, branch=self.branch)
        self.assertEqual(reg2.stock_quantity, 30)
        self.assertEqual(product2.price, 600.00)

    def test_bulk_insert_excel_validation_error(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        import openpyxl
        import io
        
        # Missing Barcode on row 2
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Barcode", "Price", "Size", "Branch Code", "Initial Stock", "Low Stock Alert"])
        ws.append(["Excel T-Shirt", "", 350.00, "M", 10001, 25, 5])
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        
        excel_file = SimpleUploadedFile(
            "bulk_invalid.xlsx",
            excel_io.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(reverse('bulk_insert'), {'csv_file': excel_file})
        self.assertEqual(response.status_code, 200)
        
        # Ensure it was rolled back and product wasn't created
        self.assertFalse(Product.objects.filter(name="Excel T-Shirt").exists())


class BranchScopedComboPriceTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.branch_a = Branch.objects.create(name="Branch A", code="10001", invoice_prefix="BA")
        self.branch_b = Branch.objects.create(name="Branch B", code="10002", invoice_prefix="BB")
        self.user = User.objects.create_user(
            username="testowner", 
            password="password123", 
            role="owner",
            active_branch=self.branch_a
        )
        self.user.branches.add(self.branch_a, self.branch_b)
        self.user.save()
        self.client.login(username="testowner", password="password123")

        self.product_a = Product.objects.create(name="Combo Shirt", barcode="777888", price=400, branch=self.branch_a)
        self.product_b = Product.objects.create(name="Combo Shirt", barcode="777888", price=400, branch=self.branch_b)
        self.registry_a = ProductRegistry.objects.create(
            branch=self.branch_a,
            product=self.product_a,
            stock_quantity=50
        )
        self.registry_b = ProductRegistry.objects.create(
            branch=self.branch_b,
            product=self.product_b,
            stock_quantity=50
        )

    def test_combos_are_branch_scoped(self):
        from core.models import ComboPrice
        combo_a = ComboPrice.objects.create(
            product=self.product_a,
            branch=self.branch_a,
            quantity=10,
            price=3000
        )

        self.assertIn(combo_a, self.registry_a.combos)
        self.assertNotIn(combo_a, self.registry_b.combos)

        combo_b = ComboPrice.objects.create(
            product=self.product_b,
            branch=self.branch_b,
            quantity=10,
            price=3500
        )
        self.assertIn(combo_b, self.registry_b.combos)
        self.assertNotIn(combo_b, self.registry_a.combos)

    def test_pos_view_uses_branch_scoped_combos(self):
        self.user.active_branch = self.branch_a
        self.user.save()

        from core.models import ComboPrice
        ComboPrice.objects.create(
            product=self.product_a,
            branch=self.branch_a,
            quantity=5,
            price=1500
        )

        ComboPrice.objects.create(
            product=self.product_b,
            branch=self.branch_b,
            quantity=5,
            price=1800
        )

        url = reverse('get_product_by_barcode') + "?barcode=777888"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data['combos']), 1)
        self.assertEqual(data['combos'][0]['price'], 1500.0)

        self.user.active_branch = self.branch_b
        self.user.save()
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data['combos']), 1)
        self.assertEqual(data['combos'][0]['price'], 1800.0)

    def test_edit_product_combos_does_not_affect_other_branch(self):
        from core.models import ComboPrice
        combo_b = ComboPrice.objects.create(
            product=self.product_b,
            branch=self.branch_b,
            quantity=10,
            price=3500
        )

        self.user.active_branch = self.branch_a
        self.user.save()

        url = reverse('product_update', args=[self.product_a.pk]) + f"?reg_id={self.registry_a.pk}"
        
        post_data = {
            'name': 'Combo Shirt',
            'barcode': '777888',
            'price': '400',
            'low_stock_threshold': '10',
            'stock_update_type': 'none',
            'stock_update_qty': '',
            'stock_update_reason': '',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)

        self.assertTrue(ComboPrice.objects.filter(pk=combo_b.pk).exists())
        self.assertEqual(ComboPrice.objects.filter(product=self.product_b, branch=self.branch_b).count(), 1)

        post_data_add = {
            'name': 'Combo Shirt',
            'barcode': '777888',
            'price': '400',
            'low_stock_threshold': '10',
            'stock_update_type': 'none',
            'stock_update_qty': '',
            'stock_update_reason': '',
            'combos-TOTAL_FORMS': '1',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
            'combos-0-quantity': '5',
            'combos-0-price': '1800',
            'combos-0-id': '',
        }
        response = self.client.post(url, post_data_add)
        self.assertEqual(response.status_code, 302)

        self.assertEqual(ComboPrice.objects.filter(product=self.product_a, branch=self.branch_a).count(), 1)
        self.assertEqual(ComboPrice.objects.filter(product=self.product_b, branch=self.branch_b).count(), 1)

        combo_a = ComboPrice.objects.get(product=self.product_a, branch=self.branch_a)
        self.assertEqual(combo_a.quantity, 5)
        self.assertEqual(combo_a.price, 1800)

    def test_barcode_and_name_can_be_edited_independently_per_branch(self):
        self.user.active_branch = self.branch_a
        self.user.save()
        
        url = reverse('product_update', args=[self.product_a.pk]) + f"?reg_id={self.registry_a.pk}"
        
        post_data = {
            'name': 'New Combo Shirt A',
            'barcode': '777888-A',
            'price': '450',
            'low_stock_threshold': '10',
            'stock_update_type': 'none',
            'stock_update_qty': '',
            'stock_update_reason': '',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        
        self.product_a.refresh_from_db()
        self.assertEqual(self.product_a.name, 'New Combo Shirt A')
        self.assertEqual(self.product_a.barcode, '777888-A')
        
        self.product_b.refresh_from_db()
        self.assertEqual(self.product_b.name, 'Combo Shirt')
        self.assertEqual(self.product_b.barcode, '777888')

    def test_add_product_duplicate_barcode_same_branch_validation_error(self):
        url = reverse('product_create')
        post_data = {
            'name': 'Duplicate Product A',
            'barcode': '777888',
            'price': '300',
            'low_stock_threshold': '10',
            'initial_branch': self.branch_a.pk,
            'initial_stock': '0',
            'stock_update_type': 'none',
            'stock_update_qty': '',
            'stock_update_reason': '',
            'combos-TOTAL_FORMS': '0',
            'combos-INITIAL_FORMS': '0',
            'combos-MIN_NUM_FORMS': '0',
            'combos-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        self.assertFormError(response.context['form'], 'barcode', "Product with barcode '777888' already exists for this branch.")


class ComboBadgeTestCase(TestCase):
    def setUp(self):
        from core.models import ComboGroup
        self.client = Client()
        self.branch_a = Branch.objects.create(name="Nellore Branch", code="10001", invoice_prefix="AN")
        self.branch_b = Branch.objects.create(name="Guntur Branch", code="10002", invoice_prefix="AG")
        self.user = User.objects.create_user(
            username="testowner", 
            password="password123", 
            role="owner",
            active_branch=self.branch_a
        )
        self.user.branches.add(self.branch_a, self.branch_b)
        self.user.save()
        self.client.login(username="testowner", password="password123")

        self.product_a = Product.objects.create(name="Product A", barcode="111111", price=500, branch=self.branch_a)
        self.registry_a = ProductRegistry.objects.create(
            branch=self.branch_a,
            product=self.product_a,
            stock_quantity=10
        )
        
        self.product_b = Product.objects.create(name="Product B", barcode="222222", price=600, branch=self.branch_b)
        self.registry_b = ProductRegistry.objects.create(
            branch=self.branch_b,
            product=self.product_b,
            stock_quantity=5
        )

        self.combo_group = ComboGroup.objects.create(name="Summer Mix & Match", is_active=True)
        self.combo_group.branches.add(self.branch_a)
        self.combo_group.products.add(self.product_a)

    def test_is_in_active_combo(self):
        # registry_a has product_a, which is in the active combo group for branch_a
        self.assertTrue(self.registry_a.is_in_active_combo)

        # registry_b has product_b, which is NOT in any combo group
        self.assertFalse(self.registry_b.is_in_active_combo)

        # If we deactivate the combo group, registry_a should be False
        self.combo_group.is_active = False
        self.combo_group.save()
        self.assertFalse(self.registry_a.is_in_active_combo)

    def test_product_list_shows_combo_badge(self):
        url = reverse('product_list') + f"?branch={self.branch_a.pk}"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<span class="badge bg-info-subtle text-info rounded-pill px-2 py-0.5 text-xxs fw-semibold"><i class="bi bi-gift-fill me-0.5"></i>Combo</span>')

    def test_product_list_combos_filter(self):
        # Product A is in an active combo group, Product B is not.
        # Request with filter=combos and branch_a should show Product A.
        url = reverse('product_list') + f"?branch={self.branch_a.pk}&filter=combos"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Product A')

        # Request with filter=combos and branch_b (which only has Product B, not in combo) should not show Product B.
        url_b = reverse('product_list') + f"?branch={self.branch_b.pk}&filter=combos"
        response_b = self.client.get(url_b)
        self.assertEqual(response_b.status_code, 200)
        self.assertNotContains(response_b, 'Product B')

    def test_product_create_preselects_branch(self):
        url = reverse('product_create') + f"?branch={self.branch_b.pk}"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        form = response.context['form']
        self.assertEqual(form.initial.get('initial_branch'), self.branch_b)

    def test_product_list_displays_branch_product_count(self):
        url = reverse('product_list')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        # Verify Nellore Branch (1 product) and Guntur Branch (1 product) show their product count
        self.assertContains(response, 'Products: 1')


class ComboListPermissionTestCase(TestCase):
    def setUp(self):
        from core.models import ComboGroup
        self.client = Client()
        self.branch_a = Branch.objects.create(name="Nellore Branch", code="10001", invoice_prefix="AN")
        self.branch_b = Branch.objects.create(name="Guntur Branch", code="10002", invoice_prefix="AG")
        
        # Owner user
        self.owner = User.objects.create_user(
            username="owner_user", 
            password="password123", 
            role="owner",
            active_branch=self.branch_a
        )
        
        # Staff user assigned ONLY to Branch A
        self.staff_a = User.objects.create_user(
            username="staff_a", 
            password="password123", 
            role="sales_staff",
            active_branch=self.branch_a
        )
        self.staff_a.branches.add(self.branch_a)
        
        # Create two combos: one for Branch A, one for Branch B
        self.combo_a = ComboGroup.objects.create(name="Nellore Deal", is_active=True)
        self.combo_a.branches.add(self.branch_a)
        
        self.combo_b = ComboGroup.objects.create(name="Guntur Deal", is_active=True)
        self.combo_b.branches.add(self.branch_b)

    def test_owner_sees_all_combos(self):
        self.client.login(username="owner_user", password="password123")
        url = reverse('combo_list')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nellore Deal")
        self.assertContains(response, "Guntur Deal")

    def test_staff_only_sees_assigned_branch_combos(self):
        self.client.login(username="staff_a", password="password123")
        url = reverse('combo_list')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nellore Deal")
        self.assertNotContains(response, "Guntur Deal")


class ComboOptimizationsTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.branch_a = Branch.objects.create(name="Nellore Branch", code="10001", invoice_prefix="AN")
        self.branch_b = Branch.objects.create(name="Guntur Branch", code="10002", invoice_prefix="AG")
        
        # User roles
        self.owner = User.objects.create_user(
            username="owner_user", 
            password="password123", 
            role="owner",
            active_branch=self.branch_a
        )
        self.staff_a = User.objects.create_user(
            username="staff_a", 
            password="password123", 
            role="sales_staff",
            active_branch=self.branch_a
        )
        self.staff_a.branches.add(self.branch_a)
        
        # Create some products
        self.prod_a = Product.objects.create(branch=self.branch_a, name="Shirt XL", barcode="11111", price=500)
        self.prod_b = Product.objects.create(branch=self.branch_b, name="Pants L", barcode="22222", price=700)

    def test_branch_products_ajax_owner(self):
        self.client.login(username="owner_user", password="password123")
        url = reverse('branch_products_ajax')
        response = self.client.get(f"{url}?branch_id={self.branch_a.id}")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data['products']), 1)
        self.assertEqual(data['products'][0]['name'], "Shirt XL")

    def test_branch_products_ajax_staff_permissions(self):
        self.client.login(username="staff_a", password="password123")
        url = reverse('branch_products_ajax')
        # Allowed for assigned branch A
        response = self.client.get(f"{url}?branch_id={self.branch_a.id}")
        self.assertEqual(response.status_code, 200)
        
        # Denied for unassigned branch B
        response = self.client.get(f"{url}?branch_id={self.branch_b.id}")
        self.assertEqual(response.status_code, 403)

    def test_combo_list_search_by_name(self):
        self.client.login(username="owner_user", password="password123")
        ComboGroup.objects.create(name="Nellore Deal", is_active=True)
        ComboGroup.objects.create(name="Guntur Deal", is_active=True)
        
        url = reverse('combo_list')
        response = self.client.get(f"{url}?q=Nellore")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nellore Deal")
        self.assertNotContains(response, "Guntur Deal")

    def test_combo_list_search_by_combo_id(self):
        self.client.login(username="owner_user", password="password123")
        combo = ComboGroup.objects.create(name="Nellore Deal", is_active=True)
        
        url = reverse('combo_list')
        response = self.client.get(f"{url}?q={combo.combo_id}")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nellore Deal")

    def test_combo_list_pagination(self):
        self.client.login(username="owner_user", password="password123")
        for i in range(12):
            ComboGroup.objects.create(name=f"Combo {i}", is_active=True)
            
        url = reverse('combo_list')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        # Verify pagination shows 10 items
        self.assertEqual(len(response.context['combos']), 10)
        self.assertTrue(response.context['page_obj'].has_other_pages())


class ComboMultiBranchTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.branch_a = Branch.objects.create(name="Nellore Branch", code="10001", invoice_prefix="AN")
        self.branch_b = Branch.objects.create(name="Guntur Branch", code="10002", invoice_prefix="AG")
        
        # Owner user
        self.owner = User.objects.create_user(
            username="owner_user", 
            password="password123", 
            role="owner",
            active_branch=self.branch_a
        )
        
        # Products with matching barcodes in both branches
        self.prod_a = Product.objects.create(branch=self.branch_a, name="Shirt XL", barcode="BAR123", price=500)
        self.prod_b = Product.objects.create(branch=self.branch_b, name="Shirt XL", barcode="BAR123", price=500)

    def test_create_combo_apply_to_all_branches(self):
        self.client.login(username="owner_user", password="password123")
        url = reverse('combo_create')
        post_data = {
            'name': 'All Branch Deal',
            'is_active': 'on',
            'products': [self.prod_a.barcode],
            'apply_to_all_branches': 'on',
            'tier_quantity[]': ['2'],
            'tier_price[]': ['900']
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302) # Redirects to list
        
        combo = ComboGroup.objects.get(name='All Branch Deal')
        # Check that both branches are assigned
        self.assertEqual(combo.branches.count(), 2)
        self.assertTrue(self.branch_a in combo.branches.all())
        self.assertTrue(self.branch_b in combo.branches.all())
        
        # Check that both products (matching barcode) are associated
        self.assertEqual(combo.products.count(), 2)
        self.assertTrue(self.prod_a in combo.products.all())
        self.assertTrue(self.prod_b in combo.products.all())

    def test_edit_combo_apply_to_specific_branches(self):
        # Create a combo initially only for branch_a
        combo = ComboGroup.objects.create(name='Nellore Only Deal', is_active=True)
        combo.branches.add(self.branch_a)
        combo.products.add(self.prod_a)
        
        self.client.login(username="owner_user", password="password123")
        url = reverse('combo_edit', kwargs={'pk': combo.pk})
        
        # Edit combo to apply to both branches
        post_data = {
            'name': 'Updated Deal',
            'is_active': 'on',
            'products': [self.prod_a.barcode],
            'selected_branches': [str(self.branch_a.id), str(self.branch_b.id)],
            'tier_quantity[]': ['2'],
            'tier_price[]': ['900']
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        
        combo.refresh_from_db()
        self.assertEqual(combo.name, 'Updated Deal')
        
        # Check that both branches are assigned
        self.assertEqual(combo.branches.count(), 2)
        self.assertTrue(self.branch_a in combo.branches.all())
        self.assertTrue(self.branch_b in combo.branches.all())
        
        # Check that both products (matching barcode) are associated
        self.assertEqual(combo.products.count(), 2)
        self.assertTrue(self.prod_a in combo.products.all())
        self.assertTrue(self.prod_b in combo.products.all())

    def test_global_products_ajax(self):
        self.client.login(username="owner_user", password="password123")
        url = reverse('global_products_ajax')
        response = self.client.get(f"{url}?q=Shirt")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data['products']), 1)
        self.assertEqual(data['products'][0]['name'], "Shirt XL")


class ComboForcedMilestoneTestCase(TestCase):
    def test_forced_milestone_application(self):
        from core.combo_views import calculate_optimal_combo_price
        # 4 items at 600 each (total 2400 without combo), combo tier is 4 items -> 2000
        # Since combo price (2000) is cheaper than regular price (2400), it should apply.
        prices = [600, 600, 600, 600]
        tiers = [(4, 2000)]
        optimal = calculate_optimal_combo_price(prices, tiers)
        self.assertEqual(optimal, 2000.0)

    def test_forced_milestone_application_with_leftover(self):
        from core.combo_views import calculate_optimal_combo_price
        # 5 items at 600 each (total 3000 without combo), combo tier is 4 items -> 2000
        # Under pro-rated flat unit price: highest achieved tier (4 items -> 2000) unit price is 500.
        # Total cost for 5 items = 5 * 500 = 2500 (which is cheaper than 3000).
        prices = [600, 600, 600, 600, 600]
        tiers = [(4, 2000)]
        optimal = calculate_optimal_combo_price(prices, tiers)
        self.assertEqual(optimal, 2500.0)

    def test_no_milestone_met(self):
        from core.combo_views import calculate_optimal_combo_price
        # 3 items at 299 each, combo tier is 4 items -> 2000
        # It must fallback to regular prices since 3 items cannot satisfy the milestone
        prices = [299, 299, 299]
        tiers = [(4, 2000)]
        optimal = calculate_optimal_combo_price(prices, tiers)
        self.assertEqual(optimal, 897.0)

    def test_better_price_rule_applied(self):
        from core.combo_views import calculate_optimal_combo_price
        # 4 items at 299 each (total 1196 without combo), combo tier is 4 items -> 2000
        # Under Better Price rule, since regular total (1196) is cheaper than combo price (2000),
        # regular total should be returned.
        prices = [299, 299, 299, 299]
        tiers = [(4, 2000)]
        optimal = calculate_optimal_combo_price(prices, tiers)
        self.assertEqual(optimal, 1196.0)


class ComboBarcodeOverlapValidationTestCase(TestCase):
    def setUp(self):
        from django.test import Client
        from core.models import ComboTier
        self.client = Client()
        self.branch = Branch.objects.create(name="Nellore Branch", code="10001", invoice_prefix="AN")
        self.owner = User.objects.create_user(
            username="owner_user", 
            password="password123", 
            role="owner",
            active_branch=self.branch
        )
        self.product = Product.objects.create(branch=self.branch, name="Shirt XL", barcode="12345", price=500)
        self.registry = ProductRegistry.objects.create(branch=self.branch, product=self.product, stock_quantity=10)

        # Create an initial active combo group
        self.combo1 = ComboGroup.objects.create(name="Nellore Deal 1", is_active=True)
        self.combo1.branches.add(self.branch)
        self.combo1.products.add(self.product)
        ComboTier.objects.create(combo_group=self.combo1, quantity=2, price=900)

    def test_create_combo_overlap_validation_error(self):
        self.client.login(username="owner_user", password="password123")
        url = reverse('combo_create')
        post_data = {
            'name': 'Nellore Deal 2',
            'is_active': 'on',
            'products': [self.product.barcode],
            'selected_branches': [str(self.branch.id)],
            'tier_quantity[]': ['3'],
            'tier_price[]': ['1200']
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nellore Deal 1")
        self.assertContains(response, "CB-0001")
        
        # Verify that combo was NOT created
        self.assertFalse(ComboGroup.objects.filter(name="Nellore Deal 2").exists())

    def test_edit_combo_overlap_validation_error(self):
        self.client.login(username="owner_user", password="password123")
        
        # Create a second active combo group but with a different barcode initially
        product2 = Product.objects.create(branch=self.branch, name="Pants L", barcode="67890", price=700)
        combo2 = ComboGroup.objects.create(name="Nellore Deal 2", is_active=True)
        combo2.branches.add(self.branch)
        combo2.products.add(product2)
        
        url = reverse('combo_edit', kwargs={'pk': combo2.pk})
        post_data = {
            'name': 'Nellore Deal 2',
            'is_active': 'on',
            'products': [self.product.barcode],  # Try to change it to the overlapping barcode
            'selected_branches': [str(self.branch.id)],
            'tier_quantity[]': ['3'],
            'tier_price[]': ['1200']
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nellore Deal 1")
        self.assertContains(response, "CB-0001")
        
        # Verify that product2 was NOT updated to the overlapping barcode in combo2
        combo2.refresh_from_db()
        self.assertNotIn(self.product, combo2.products.all())







