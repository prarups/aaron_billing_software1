from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Product, Branch, ProductRegistry, StockTransaction, StockAdjustment, ComboPrice
from django import forms
from .forms import ProductForm
from django.http import JsonResponse
from billing.models import Bill


@login_required
def product_list(request):
    if request.user.role == 'sales_staff':
        return redirect('dashboard')
    
    # Auto-initialize active branch if None
    if not request.user.active_branch:
        accessible = request.user.get_accessible_branches()
        if accessible.exists():
            request.user.active_branch = accessible.first()
            request.user.save()
    
    from django.db.models import Q, Count, Sum, F, DecimalField, ExpressionWrapper
    from django.core.paginator import Paginator

    q = request.GET.get('q', '')
    selected_branch = request.GET.get('branch', '')
    active_filter = request.GET.get('filter', '')
    
    # Base query
    accessible_branches = request.user.get_accessible_branches()
    registrations = ProductRegistry.objects.select_related('product', 'branch').prefetch_related(
        'product__combos',
        'product__combo_groups',
        'product__combo_groups__branches'
    ).filter(branch__in=accessible_branches)

    if q:
        registrations = registrations.filter(
            Q(product__name__icontains=q) | Q(product__barcode__icontains=q)
        )

    if selected_branch:
        registrations = registrations.filter(branch_id=selected_branch)
    
    if active_filter == 'low_stock':
        registrations = registrations.filter(stock_quantity__lte=F('low_stock_threshold'))
    elif active_filter == 'zero_stock':
        registrations = registrations.filter(stock_quantity=0)
    elif active_filter == 'combos':
        registrations = registrations.filter(
            product__combo_groups__is_active=True,
            product__combo_groups__branches=F('branch')
        ).distinct()
        
    registrations = registrations.order_by('product__name')

    # Stats
    stats = registrations.aggregate(
        product_count=Count('product', distinct=True),
        registration_count=Count('id'),
        total_value=Sum(
            ExpressionWrapper(F('stock_quantity') * F('product__price'), output_field=DecimalField())
        )
    )

    # Pagination
    paginator = Paginator(registrations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if active_filter == 'low_stock':
        branches = accessible_branches.annotate(
            low_stock_count=Count(
                'productregistry',
                filter=Q(productregistry__stock_quantity__lte=F('productregistry__low_stock_threshold'))
            )
        ).filter(low_stock_count__gt=0)
    else:
        branches = accessible_branches.annotate(product_count=Count('productregistry'))

    return render(request, 'core/product_list.html', {
        'page_obj': page_obj,
        'stats': stats,
        'q': q,
        'branches': branches,
        'selected_branch': selected_branch,
        'active_filter': active_filter,
    })

@login_required
def export_products_csv(request):
    if request.user.role == 'sales_staff':
        return redirect('dashboard')
        
    import csv
    from django.http import HttpResponse

    q = request.GET.get('q', '')
    selected_branch = request.GET.get('branch', '')
    
    accessible_branches = request.user.get_accessible_branches()
    registrations = ProductRegistry.objects.select_related('product', 'branch').prefetch_related(
        'product__combo_groups',
        'product__combo_groups__branches'
    ).filter(branch__in=accessible_branches)

    if q:
        from django.db.models import Q
        registrations = registrations.filter(
            Q(product__name__icontains=q) | Q(product__barcode__icontains=q)
        )
    
    if selected_branch:
        registrations = registrations.filter(branch_id=selected_branch)

    registrations = registrations.order_by('-created_at')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'

    writer = csv.writer(response)
    writer.writerow(['Branch Name', 'Branch Code', 'Product Name', 'Barcode', 'Price', 'Combo', 'Stock', 'Low Stock Level', 'Registered On'])

    for reg in registrations:
        date_str = reg.created_at.strftime("%Y-%m-%d %H:%M")
        combo_str = 'Yes' if reg.is_in_active_combo else '-'
        writer.writerow([
            reg.branch.name, 
            reg.branch.code or '',
            reg.product.name, 
            reg.product.barcode, 
            str(int(reg.product.price)),
            combo_str,
            reg.stock_quantity,
            reg.low_stock_threshold,
            date_str
        ])

    return response


@login_required
def product_create(request):
    if not (request.user.role == 'owner' or request.user.has_product_rights):
        messages.error(request, "Permission denied. You do not have product edit rights.")
        return redirect('product_list')

    next_url = request.GET.get('next') or request.POST.get('next') or request.META.get('HTTP_REFERER') or ''
    if '/products/' not in next_url:
        from django.urls import reverse
        next_url = reverse('product_list')
    
    # Auto-initialize active branch if None
    if not request.user.active_branch:
        accessible = request.user.get_accessible_branches()
        if accessible.exists():
            request.user.active_branch = accessible.first()
            request.user.save()

    if request.method == 'POST':
        # Determine branch first to scope the product's uniqueness check
        initial_branch_id = request.POST.get('initial_branch')
        if initial_branch_id:
            try:
                branch = Branch.objects.get(id=initial_branch_id)
            except Branch.DoesNotExist:
                branch = request.user.active_branch
        else:
            branch = request.user.active_branch

        product_instance = Product(branch=branch)
        form = ProductForm(request.POST, instance=product_instance)
        from .forms import ComboPriceFormSet
        combo_formset = ComboPriceFormSet(request.POST)
        
        has_combo_formset = 'combos-TOTAL_FORMS' in request.POST
        if form.is_valid() and (not has_combo_formset or combo_formset.is_valid()):
            product = form.save()
            if has_combo_formset:
                combo_formset.instance = product
                combos = combo_formset.save(commit=False)
                for combo in combos:
                    combo.branch = branch
                    combo.save()
                for obj in combo_formset.deleted_objects:
                    obj.delete()
            
            initial_stock = form.cleaned_data.get('initial_stock') or 0
            low_threshold = form.cleaned_data.get('low_stock_threshold') or 10
            
            if branch:
                ProductRegistry.objects.create(
                    branch=branch,
                    product=product,
                    stock_quantity=initial_stock,
                    low_stock_threshold=low_threshold,
                )
                # Manual stock input maps to opening stock; no StockTransaction is recorded for this initial quantity.
                return redirect(next_url)
        else:
            pass
    else:
        initial_branch = request.user.active_branch
        branch_id = request.GET.get('branch')
        if branch_id:
            try:
                initial_branch = Branch.objects.get(id=branch_id)
            except Branch.DoesNotExist:
                pass
        form = ProductForm(initial={
            'initial_branch': initial_branch,
            'initial_stock': 0,
            'low_stock_threshold': 10
        })
        from .forms import ComboPriceFormSet
        combo_formset = ComboPriceFormSet()
    return render(request, 'core/product_form.html', {'form': form, 'combo_formset': combo_formset, 'action': 'Add New', 'is_admin': True, 'next_url': next_url})

@login_required
def product_update(request, pk):
    is_admin = (request.user.role == 'owner' or request.user.has_product_rights)

    product = get_object_or_404(Product, pk=pk)
    
    next_url = request.GET.get('next') or request.POST.get('next') or request.META.get('HTTP_REFERER') or ''
    if '/products/' not in next_url:
        from django.urls import reverse
        next_url = reverse('product_list')
    
    # For updating, we might want to edit a specific registration
    reg_id = request.GET.get('reg_id')
    registration = None
    if reg_id:
        registration = get_object_or_404(ProductRegistry, id=reg_id, product=product)

    current_damaged_qty = 0
    if registration:
        from django.db.models import Sum
        current_damaged_qty = StockTransaction.objects.filter(
            product=product,
            branch=registration.branch,
            transaction_type='DMG'
        ).aggregate(t=Sum('quantity'))['t'] or 0

    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        from .forms import ComboPriceFormSet
        
        branch = registration.branch if registration else request.user.active_branch
        combo_formset = ComboPriceFormSet(
            request.POST, 
            instance=product, 
            queryset=ComboPrice.objects.filter(product=product, branch=branch)
        )
        
        has_combo_formset = 'combos-TOTAL_FORMS' in request.POST
        if is_admin and form.is_valid() and (not has_combo_formset or combo_formset.is_valid()):
            product = form.save()
            if has_combo_formset:
                combos = combo_formset.save(commit=False)
                for combo in combos:
                    combo.branch = branch
                    combo.save()
                for obj in combo_formset.deleted_objects:
                    obj.delete()
        # For non-admin users, we still want to process stock updates (e.g., damage),
        # so we no longer return early here. The form validation errors (if any) will be
        # displayed after the stock processing block.

        
        # If we are editing stock for a specific registration
        if registration:
            old_stock = registration.stock_quantity
            new_low = int(request.POST.get('low_stock_threshold', registration.low_stock_threshold))
            update_type = request.POST.get('stock_update_type', 'none')
            update_qty_str = request.POST.get('stock_update_qty', '').strip()
            update_reason = request.POST.get('stock_update_reason', '').strip()
            
            has_error = False
            stock_update_error = None
            if new_low < 0:
                messages.error(request, "Low stock threshold cannot be negative.")
                has_error = True
            
            update_qty = 0
            if update_type in ['add', 'damage', 'correction', 'correct_damage']:
                # Non-admin users can only use damage and correct_damage
                if not is_admin and update_type in ('add', 'correction'):
                    messages.error(request, "You do not have permission to perform this action.")
                    has_error = True
                elif not update_qty_str:
                    stock_update_error = "Please enter a quantity for the stock update."
                    has_error = True
                else:
                    try:
                        update_qty = int(float(update_qty_str))
                        if update_qty < 0:
                            raise ValueError()
                    except ValueError:
                        stock_update_error = "Stock update quantity must be a non-negative whole number."
                        has_error = True
            
            new_stock = old_stock
            if not has_error:
                if update_type == 'add':
                    if update_qty > 0:
                        new_stock = old_stock + update_qty
                        # Update damaged quantity in registry
                        registration.stock_quantity = new_stock
                        registration.save()
                        StockTransaction.objects.create(
                            product=product,
                            branch=registration.branch,
                            transaction_type='IN',
                            quantity=update_qty,
                            reference=f"Stock Added: {update_reason}" if update_reason else "Stock Added",
                            user=request.user
                        )
                        messages.success(request, f"Added {update_qty} items to stock.")
                elif update_type == 'damage':
                    if update_qty > 0:
                        new_stock = old_stock - update_qty
                        if new_stock < 0:
                            messages.error(request, f"Cannot log {update_qty} damaged items. Current stock is only {old_stock}.")
                            has_error = True
                        else:
                            # Update registry quantities
                            registration.stock_quantity = new_stock
                            registration.damaged_qty = (registration.damaged_qty or 0) + update_qty
                            registration.save()
                            StockTransaction.objects.create(
                                product=product,
                                branch=registration.branch,
                                transaction_type='DMG',
                                quantity=update_qty,
                                reference=f"Damage Logged: {update_reason}" if update_reason else "Damage Logged",
                                user=request.user
                            )
                            
                            from core.models import StockAdjustment
                            from django.db.models import Sum
                            txns = StockTransaction.objects.filter(product=product, branch=registration.branch)
                            period_in = txns.filter(transaction_type__in=['IN', 'ADJ']).aggregate(t=Sum('quantity'))['t'] or 0
                            period_out = txns.filter(transaction_type='OUT').aggregate(t=Sum('quantity'))['t'] or 0
                            opening_balance = old_stock - period_in + period_out
                            
                            StockAdjustment.objects.create(
                                product=product,
                                branch=registration.branch,
                                opening_balance=opening_balance,
                                stock_in=period_in,
                                stock_out=period_out,
                                correction_amount=-update_qty,
                                closing_stock=new_stock,
                                is_in_stock=(new_stock > 0),
                                reason=f"Damage Logged: {update_reason}" if update_reason else "Damage Logged",
                                user=request.user
                            )
                            messages.success(request, f"Logged {update_qty} damaged items.")
                elif update_type == 'correction':
                    if update_qty_str:
                        new_stock = update_qty
                        if new_stock < 0:
                            messages.error(request, "Correction closing stock cannot be negative.")
                            has_error = True
                        else:
                            diff = new_stock - old_stock
                            if diff != 0:
                                import datetime
                                from django.utils import timezone
                                
                                # Find if there is an IN transaction today for this product/branch
                                today = timezone.now().date()
                                start_datetime = timezone.make_aware(datetime.datetime.combine(today, datetime.time.min))
                                end_datetime = timezone.make_aware(datetime.datetime.combine(today, datetime.time.max))
                                today_in_txn = StockTransaction.objects.filter(
                                    product=product,
                                    branch=registration.branch,
                                    transaction_type='IN',
                                    created_at__gte=start_datetime,
                                    created_at__lte=end_datetime
                                ).first()
                                
                                if diff > 0:
                                    # Positive correction: check today's DMG txn first (to reverse wrong damage entries)
                                    today_dmg_txn = StockTransaction.objects.filter(
                                        product=product,
                                        branch=registration.branch,
                                        transaction_type='DMG',
                                        created_at__gte=start_datetime,
                                        created_at__lte=end_datetime
                                    ).first()
                                    
                                    remainder_diff = diff
                                    if today_dmg_txn:
                                        if today_dmg_txn.quantity >= remainder_diff:
                                            today_dmg_txn.quantity -= remainder_diff
                                            if today_dmg_txn.quantity == 0:
                                                today_dmg_txn.delete()
                                            else:
                                                today_dmg_txn.reference = f"{today_dmg_txn.reference} - Corrected: {update_reason}" if update_reason else f"{today_dmg_txn.reference} - Corrected (-{remainder_diff})"
                                                today_dmg_txn.save()
                                            remainder_diff = 0
                                        else:
                                            remainder_diff -= today_dmg_txn.quantity
                                            today_dmg_txn.delete()
                                            
                                    if remainder_diff > 0:
                                        # Positive correction: add to today's IN txn if present, else create new IN txn
                                        if today_in_txn:
                                            today_in_txn.quantity += remainder_diff
                                            today_in_txn.reference = f"{today_in_txn.reference} + Correction: {update_reason}" if update_reason else f"{today_in_txn.reference} + Correction (+{remainder_diff})"
                                            today_in_txn.save()
                                        else:
                                            StockTransaction.objects.create(
                                                product=product,
                                                branch=registration.branch,
                                                transaction_type='IN',
                                                quantity=remainder_diff,
                                                reference=f"Correction: {update_reason}" if update_reason else f"Correction (+{remainder_diff})",
                                                user=request.user
                                            )
                                else:
                                    # Negative correction: reduce from today's IN txn if present, else create negative ADJ txn
                                    abs_diff = abs(diff)
                                    if today_in_txn:
                                        if today_in_txn.quantity >= abs_diff:
                                            # Can fully absorb in today's IN txn
                                            today_in_txn.quantity -= abs_diff
                                            if today_in_txn.quantity == 0:
                                                today_in_txn.delete()
                                            else:
                                                today_in_txn.reference = f"{today_in_txn.reference} - Correction: {update_reason}" if update_reason else f"{today_in_txn.reference} - Correction (-{abs_diff})"
                                                today_in_txn.save()
                                        else:
                                            # Part is absorbed, remainder is ADJ
                                            remaining = abs_diff - today_in_txn.quantity
                                            today_in_txn.delete()
                                            StockTransaction.objects.create(
                                                product=product,
                                                branch=registration.branch,
                                                transaction_type='ADJ',
                                                quantity=-remaining,
                                                reference=f"Correction: {update_reason}" if update_reason else f"Correction (-{remaining})",
                                                user=request.user
                                            )
                                    else:
                                        # No IN txn today, create negative ADJ txn
                                        StockTransaction.objects.create(
                                            product=product,
                                            branch=registration.branch,
                                            transaction_type='ADJ',
                                            quantity=diff,
                                            reference=f"Correction ({diff}): {update_reason}" if update_reason else f"Correction ({diff})",
                                            user=request.user
                                        )
                                
                                from core.models import StockAdjustment
                                from django.db.models import Sum
                                txns = StockTransaction.objects.filter(product=product, branch=registration.branch)
                                period_in = txns.filter(transaction_type__in=['IN', 'ADJ']).aggregate(t=Sum('quantity'))['t'] or 0
                                period_out = txns.filter(transaction_type='OUT').aggregate(t=Sum('quantity'))['t'] or 0
                                opening_balance = old_stock - period_in + period_out
                                
                                StockAdjustment.objects.create(
                                    product=product,
                                    branch=registration.branch,
                                    opening_balance=opening_balance,
                                    stock_in=period_in,
                                    stock_out=period_out,
                                    correction_amount=diff,
                                    closing_stock=new_stock,
                                    is_in_stock=(new_stock > 0),
                                    reason=f"Correction: {update_reason}" if update_reason else "Wrong Entry Correction",
                                    user=request.user
                                )
                                messages.success(request, f"Corrected stock to {new_stock} (Adjustment: {'+' if diff >= 0 else ''}{diff}).")
                elif update_type == 'correct_damage':
                    if update_qty_str:
                        new_damaged_total = update_qty
                        if new_damaged_total < 0:
                            messages.error(request, "Total damaged stock cannot be negative.")
                            has_error = True
                        else:
                            from django.db.models import Sum
                            old_damaged_total = StockTransaction.objects.filter(
                                product=product,
                                branch=registration.branch,
                                transaction_type='DMG'
                            ).aggregate(t=Sum('quantity'))['t'] or 0
                            
                            diff = new_damaged_total - old_damaged_total
                            if diff != 0:
                                new_stock = old_stock - diff
                                if new_stock < 0:
                                    messages.error(request, f"Correction would result in negative sellable stock ({new_stock}). Please enter a valid correction.")
                                    has_error = True
                                else:
                                    import datetime
                                    from django.utils import timezone
                                    
                                    # Find if there is a DMG transaction today for this product/branch
                                    today = timezone.now().date()
                                    start_datetime = timezone.make_aware(datetime.datetime.combine(today, datetime.time.min))
                                    end_datetime = timezone.make_aware(datetime.datetime.combine(today, datetime.time.max))
                                    today_dmg_txn = StockTransaction.objects.filter(
                                        product=product,
                                        branch=registration.branch,
                                        transaction_type='DMG',
                                        created_at__gte=start_datetime,
                                        created_at__lte=end_datetime
                                    ).first()
                                    
                                    if diff < 0:
                                        # We are reducing damage (diff is negative).
                                        # We check today's DMG txn first.
                                        abs_diff = abs(diff)
                                        if today_dmg_txn:
                                            if today_dmg_txn.quantity >= abs_diff:
                                                today_dmg_txn.quantity -= abs_diff
                                                if today_dmg_txn.quantity == 0:
                                                    today_dmg_txn.delete()
                                                else:
                                                    today_dmg_txn.reference = f"{today_dmg_txn.reference} - Corrected: {update_reason}" if update_reason else f"{today_dmg_txn.reference} - Corrected (-{abs_diff})"
                                                    today_dmg_txn.save()
                                            else:
                                                remaining = abs_diff - today_dmg_txn.quantity
                                                today_dmg_txn.delete()
                                                StockTransaction.objects.create(
                                                    product=product,
                                                    branch=registration.branch,
                                                    transaction_type='DMG',
                                                    quantity=-remaining,
                                                    reference=f"Damage Correction: {update_reason}" if update_reason else f"Damage Correction (-{remaining})",
                                                    user=request.user
                                                )
                                        else:
                                            # No today's DMG transaction, log a negative DMG transaction
                                            StockTransaction.objects.create(
                                                product=product,
                                                branch=registration.branch,
                                                transaction_type='DMG',
                                                quantity=diff,
                                                reference=f"Damage Correction ({diff}): {update_reason}" if update_reason else f"Damage Correction ({diff})",
                                                user=request.user
                                            )
                                    else:
                                        # We are increasing damage (diff is positive).
                                        # We check today's DMG txn first.
                                        if today_dmg_txn:
                                            today_dmg_txn.quantity += diff
                                            today_dmg_txn.reference = f"{today_dmg_txn.reference} + Correction: {update_reason}" if update_reason else f"{today_dmg_txn.reference} + Correction (+{diff})"
                                            today_dmg_txn.save()
                                        else:
                                            StockTransaction.objects.create(
                                                product=product,
                                                branch=registration.branch,
                                                transaction_type='DMG',
                                                quantity=diff,
                                                reference=f"Damage Correction (+{diff}): {update_reason}" if update_reason else f"Damage Correction (+{diff})",
                                                user=request.user
                                            )
                                    
                                    # Log StockAdjustment audit record
                                    from core.models import StockAdjustment
                                    txns = StockTransaction.objects.filter(product=product, branch=registration.branch)
                                    period_in = txns.filter(transaction_type__in=['IN', 'ADJ']).aggregate(t=Sum('quantity'))['t'] or 0
                                    period_out = txns.filter(transaction_type='OUT').aggregate(t=Sum('quantity'))['t'] or 0
                                    # StockAdjustment represents sellable stock correction, which is -diff
                                    opening_balance = old_stock - period_in + period_out
                                    
                                    StockAdjustment.objects.create(
                                        product=product,
                                        branch=registration.branch,
                                        opening_balance=opening_balance,
                                        stock_in=period_in,
                                        stock_out=period_out,
                                        correction_amount=-diff,
                                        closing_stock=new_stock,
                                        is_in_stock=(new_stock > 0),
                                        reason=f"Damage Correction: {update_reason}" if update_reason else "Damage Correction",
                                        user=request.user
                                    )
                                    
                                    # Set new stock and damaged_qty
                                    new_stock = new_stock
                                    registration.damaged_qty = new_damaged_total
                                    messages.success(request, f"Corrected total damaged stock to {new_damaged_total} (Stock adjusted to {new_stock}).")

            if has_error:
                # Keep the bound form and formset containing user inputs and errors
                return render(request, 'core/product_form.html', {
                    'form': form, 
                    'combo_formset': combo_formset,
                    'action': 'Edit', 
                    'registration': registration,
                    'current_damaged_qty': current_damaged_qty,
                    'stock_update_error': stock_update_error,
                    'is_admin': is_admin,
                    'next_url': next_url
                })
            
            registration.stock_quantity = new_stock
            registration.low_stock_threshold = new_low
            registration.save()
        
        return redirect(next_url)
    else:
        initial_data = {}
        if registration:
            initial_data = {
                'initial_branch': registration.branch,
                'initial_stock': registration.stock_quantity,
                'low_stock_threshold': registration.low_stock_threshold
            }
        form = ProductForm(instance=product, initial=initial_data)
        from .forms import ComboPriceFormSet
        
        branch = registration.branch if registration else request.user.active_branch
        combo_formset = ComboPriceFormSet(
            instance=product,
            queryset=ComboPrice.objects.filter(product=product, branch=branch)
        )
        
    return render(request, 'core/product_form.html', {
        'form': form, 
        'combo_formset': combo_formset,
        'action': 'Edit', 
        'registration': registration,
        'current_damaged_qty': current_damaged_qty,
        'is_admin': is_admin,
        'next_url': next_url
    })

@login_required
def update_product_stock_ajax(request, pk):
    # This pk is the ProductRegistry ID
    if not (request.user.role == 'owner' or request.user.has_product_rights):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_stock = data.get('stock')
            if new_stock is not None:
                if int(new_stock) < 0:
                    return JsonResponse({'error': 'Stock cannot be negative.'}, status=400)
                reg = get_object_or_404(ProductRegistry, pk=pk)
                old_stock = reg.stock_quantity
                reg.stock_quantity = int(new_stock)
                reg.save()
                
                if reg.stock_quantity != old_stock:
                    diff = reg.stock_quantity - old_stock
                    StockTransaction.objects.create(
                        product=reg.product,
                        branch=reg.branch,
                        transaction_type='IN' if diff > 0 else 'OUT',
                        quantity=abs(diff),
                        reference='Ajax Update',
                        user=request.user
                    )
                    
                return JsonResponse({'success': True, 'stock': reg.stock_quantity})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
            
    return JsonResponse({'error': 'Invalid request'}, status=405)

from django.http import JsonResponse
import json

@login_required
def update_product_price_ajax(request, pk):
    if not (request.user.role == 'owner' or request.user.has_product_rights):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_price = data.get('price')
            if new_price is not None:
                product = get_object_or_404(Product, pk=pk)
                product.price = new_price
                product.save()
                return JsonResponse({'success': True, 'price': float(product.price)})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request'}, status=405)

import csv
from django.db import transaction
from django.core.exceptions import ValidationError
import io
@login_required
def bulk_insert(request):
    if not (request.user.role == 'owner' or request.user.has_product_rights):
        messages.error(request, "Permission denied. You do not have product edit rights.")
        return redirect('product_list')

    
    branches = request.user.get_accessible_branches()
    results = None
    
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            messages.error(request, 'Please upload a CSV or Excel file.')
            return render(request, 'core/bulk_insert.html', {'branches': branches})
        
        file_name = csv_file.name.lower()
        if not (file_name.endswith('.csv') or file_name.endswith('.xlsx')):
            messages.error(request, 'File must be a .csv or .xlsx file.')
            return render(request, 'core/bulk_insert.html', {'branches': branches})
        
        try:
            reader_data = []
            if file_name.endswith('.xlsx'):
                from openpyxl import load_workbook
                wb = load_workbook(filename=io.BytesIO(csv_file.read()), data_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    raise ValidationError("The Excel file is empty.")
                
                headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
                
                for row in rows[1:]:
                    # Skip completely empty rows
                    if all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if header:
                            val = row[i] if i < len(row) else None
                            if val is not None:
                                if isinstance(val, float) and val.is_integer():
                                    val_str = str(int(val))
                                else:
                                    val_str = str(val).strip()
                            else:
                                val_str = ''
                            row_dict[header] = val_str
                    reader_data.append(row_dict)
                
                if not reader_data:
                    raise ValidationError("The Excel file has no data rows.")
            else:
                decoded_file = csv_file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded_file))
                reader_data = list(reader)
            
            success_count = 0
            error_count = 0
            errors = []
            seen_barcodes = set()
            seen_names = set()
            
            with transaction.atomic():
                for row_num, row in enumerate(reader_data, start=2):
                    try:
                        # Support both template formats:
                        # New template: Name, Barcode, Price, Size, Branch Code, Initial Stock, Low Stock Alert
                        # Old template: Branch, Product Name, Barcode, Size, Price, Stock, Low Stock Level
                        product_name = (row.get('Name') or row.get('Product Name') or '').strip()
                        barcode = (row.get('Barcode') or '').strip()
                        size = (row.get('HSN') or row.get('Size') or '').strip()
                        price = (row.get('Price') or '').strip()
                        stock = (row.get('Initial Stock') or row.get('Stock') or '').strip()
                        low_stock = (row.get('Low Stock Alert') or row.get('Low Stock Level') or '10').strip()
                        branch_code = (row.get('Branch Code') or '').strip()
                        branch_name = (row.get('Branch') or '').strip()
                        
                        # 1. Enforce Mandatory Fields
                        if not barcode:
                            raise ValidationError(f"Row {row_num}: Barcode is a mandatory field and cannot be empty.")
                        if not branch_code and not branch_name:
                            raise ValidationError(f"Row {row_num}: Branch Code / Branch is a mandatory field and cannot be empty.")
                        if not stock:
                            raise ValidationError(f"Row {row_num}: Initial Stock / Stock is a mandatory field and cannot be empty.")
                        if not price:
                            raise ValidationError(f"Row {row_num}: Price is a mandatory field and cannot be empty.")
                        
                        try:
                            parsed_price = float(price)
                            if parsed_price < 0:
                                raise ValueError()
                        except ValueError:
                            raise ValidationError(f"Row {row_num}: Price must be a non-negative number.")
                        
                        # Find or create branch (by code first, then by name)
                        branch = None
                        if branch_code:
                            try:
                                branch = Branch.objects.get(code=int(float(branch_code)))
                            except (Branch.DoesNotExist, ValueError):
                                raise ValidationError(f"Row {row_num}: Branch with code '{branch_code}' not found.")
                        elif branch_name:
                            branch, _ = Branch.objects.get_or_create(name=branch_name)
                        
                        branch_id = branch.id if branch else None
                        
                        # 2. Check for Duplicates in the uploaded file
                        barcode_key = (branch_id, barcode.lower())
                        if barcode_key in seen_barcodes:
                            raise ValidationError(f"Row {row_num}: Duplicate barcode '{barcode}' found for branch '{branch.name if branch else ''}' in the file.")
                        seen_barcodes.add(barcode_key)
                        
                        # 3. Check for Duplicates in the Database for this branch
                        existing_product = Product.objects.filter(branch=branch, barcode__iexact=barcode).first()
                        if existing_product:
                            raise ValidationError(f"Row {row_num}: Product with barcode '{barcode}' already exists for branch '{branch.name}' in the database.")
                        
                        # Create product scoped to this branch
                        product = Product.objects.create(
                            barcode=barcode,
                            name=product_name,
                            price=parsed_price,
                            size=size if size else '',
                            branch=branch
                        )
                        
                        # Create registry entry if branch is provided
                        if branch:
                            try:
                                parsed_stock = int(float(stock))
                                if parsed_stock < 0:
                                    raise ValueError()
                            except ValueError:
                                raise ValidationError(f"Row {row_num}: Initial Stock must be a non-negative integer.")
                            
                            try:
                                parsed_low_stock = max(0, int(float(low_stock)) if low_stock else 10)
                            except ValueError:
                                raise ValidationError(f"Row {row_num}: Low Stock Alert must be an integer.")
                            
                            reg = ProductRegistry.objects.create(
                                branch=branch,
                                product=product,
                                stock_quantity=parsed_stock,
                                low_stock_threshold=parsed_low_stock
                            )
                            
                            if reg.stock_quantity > 0:
                                StockTransaction.objects.create(
                                    product=product,
                                    branch=branch,
                                    transaction_type='IN',
                                    quantity=reg.stock_quantity,
                                    reference='Bulk Insert',
                                    user=request.user
                                )
                        
                        success_count += 1
                        
                    except ValidationError as ve:
                        errors.append(str(ve))
                        error_count += 1
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        error_count += 1
                
                if error_count > 0:
                    # Raise to trigger transaction rollback and show detailed errors
                    raise ValidationError(f"Bulk insert failed with {error_count} error(s). Details: {', '.join(errors)}")
            results = {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            }
            
        except ValidationError as ve:
            messages.error(request, str(ve))
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
    
    return render(request, 'core/bulk_insert.html', {
        'branches': branches,
        'results': results,
    })

@login_required
def download_bulk_template(request):
    # Updated bulk insert template header to include Branch Code
    header = ["Product Name","Barcode","Price","HSN","Branch Code","Initial Stock","Low Stock Alert"]
    # This endpoint will generate a CSV template for bulk insert
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bulk_insert_template.csv"'
    writer = csv.writer(response)
    writer.writerow(header)
    writer.writerow(['Example Product', 'EX001', '299.00', '8409', 'BR001', '50', '10'])
    writer.writerow(['Another Product', 'EX002', '149.00', '8409', 'BR001', '100', '15'])
    
    return response

@login_required
def stock_pivot_report(request):
    """Report showing products, their stock movement (Op, In, Out, Cl) and branch-wise closing stock."""
    if request.user.role == 'sales_staff':
        return redirect('dashboard')
    
    from django.db.models import Sum, Q, F, Count
    import datetime
    from django.utils import timezone
    from django.core.paginator import Paginator
    
    q = request.GET.get('q', '').strip()
    from_date_str = request.GET.get('from_date', '')
    to_date_str = request.GET.get('to_date', '')
    
    today = timezone.now().date()
    start_date = today
    end_date = today
    
    if from_date_str:
        try:
            start_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
            
    if to_date_str:
        try:
            end_date = datetime.datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
        
    accessible_branches = request.user.get_accessible_branches().annotate(product_count=Count('productregistry')).order_by('name')
    branch_id = request.GET.get('branch', '').strip()
    
    selected_branch = None
    if branch_id:
        try:
            selected_branch = accessible_branches.get(id=int(branch_id))
        except (ValueError, Branch.DoesNotExist):
            pass

    if selected_branch:
        active_branches = [selected_branch]
    else:
        active_branches = list(accessible_branches)

    if q:
        products = Product.objects.filter(
            Q(barcode__icontains=q) | Q(name__icontains=q),
            branches__in=active_branches
        ).distinct().order_by('name')
    else:
        products = Product.objects.filter(branches__in=active_branches).distinct().order_by('name')

    # Pagination: 10 products per page
    paginator = Paginator(products, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    print("DEBUG PAGINATION: total products:", paginator.count, "has_other:", page_obj.has_other_pages())

    if request.GET.get('test_error'):
        raise Exception(f"Pagination triggered: {paginator.count} items, returning {len(page_obj.object_list)}")

    # 1. Fetch current stock via registries (restricted to active branches)
    registries = ProductRegistry.objects.filter(branch__in=active_branches).values('id', 'product_id', 'branch_id', 'stock_quantity')
    stock_map = {(r['product_id'], r['branch_id']): r['stock_quantity'] for r in registries}
    reg_id_map = {(r['product_id'], r['branch_id']): r['id'] for r in registries}
    
    # 2. Fetch stock movements (Op, In, Ret, Out, Dmg, Ret Dmg, Cl)
    start_datetime = timezone.make_aware(datetime.datetime.combine(start_date, datetime.time.min))
    end_datetime = timezone.make_aware(datetime.datetime.combine(end_date, datetime.time.max))

    # Period transactions using conditional aggregation
    period_txns = StockTransaction.objects.filter(
        created_at__range=(start_datetime, end_datetime),
        branch__in=active_branches
    ).values('product_id', 'branch_id').annotate(
        in_std=Sum('quantity', filter=Q(transaction_type='IN') & ~Q(reference__startswith='Return #')),
        ret_good=Sum('quantity', filter=Q(transaction_type='IN') & Q(reference__startswith='Return #')),
        out_gross=Sum('quantity', filter=Q(transaction_type='OUT')),
        dmg_std=Sum('quantity', filter=Q(transaction_type='DMG') & ~Q(reference__startswith='Return #')),
        ret_dmg=Sum('quantity', filter=Q(transaction_type='DMG') & Q(reference__startswith='Return #')),
        adj=Sum('quantity', filter=Q(transaction_type='ADJ')),
    )
    period_map = {(t['product_id'], t['branch_id']): t for t in period_txns}

    # All-time transactions using conditional aggregation
    all_time_txns = StockTransaction.objects.filter(
        branch__in=active_branches
    ).values('product_id', 'branch_id').annotate(
        out_gross=Sum('quantity', filter=Q(transaction_type='OUT')),
        ret_good=Sum('quantity', filter=Q(transaction_type='IN') & Q(reference__startswith='Return #')),
        ret_dmg=Sum('quantity', filter=Q(transaction_type='DMG') & Q(reference__startswith='Return #')),
        dmg_std=Sum('quantity', filter=Q(transaction_type='DMG') & ~Q(reference__startswith='Return #')),
        adj=Sum('quantity', filter=Q(transaction_type='ADJ')),
    )
    all_time_map = {(t['product_id'], t['branch_id']): t for t in all_time_txns}
        
    future_map = {}
    if end_date < today:
        future_txns = StockTransaction.objects.filter(
            created_at__gt=end_datetime,
            branch__in=active_branches
        ).values('product_id', 'branch_id').annotate(
            in_std=Sum('quantity', filter=Q(transaction_type='IN') & ~Q(reference__startswith='Return #')),
            ret_good=Sum('quantity', filter=Q(transaction_type='IN') & Q(reference__startswith='Return #')),
            out_gross=Sum('quantity', filter=Q(transaction_type='OUT')),
            dmg_std=Sum('quantity', filter=Q(transaction_type='DMG') & ~Q(reference__startswith='Return #')),
            adj=Sum('quantity', filter=Q(transaction_type='ADJ')),
        )
        future_map = {(t['product_id'], t['branch_id']): t for t in future_txns}

    report_data = []
    
    for p in page_obj:
        item = {
            'product': p,
            'total_op': 0,
            'total_in': 0,
            'total_ret': 0,
            'total_out': 0,
            'total_adj_plus': 0,
            'total_adj_minus': 0,
            'total_dmg': 0,
            'total_ret_dmg': 0,
            'total_cl': 0,
            'total_rec': 0,
            'total_all_time_out': 0,
            'total_all_time_dmg': 0,
            'branch_stocks': []
        }
        
        for b in active_branches:
            curr_stock = stock_map.get((p.id, b.id), 0)
            
            p_tx = period_map.get((p.id, b.id), {})
            period_in_std = p_tx.get('in_std') or 0
            period_ret = p_tx.get('ret_good') or 0
            period_out_gross = p_tx.get('out_gross') or 0
            period_dmg_std = p_tx.get('dmg_std') or 0
            period_ret_dmg = p_tx.get('ret_dmg') or 0
            period_adj = p_tx.get('adj') or 0
            
            period_adj_plus = period_adj if period_adj > 0 else 0
            period_adj_minus = abs(period_adj) if period_adj < 0 else 0
            
            # All-time aggregates for all-time columns
            a_tx = all_time_map.get((p.id, b.id), {})
            all_time_out_gross = a_tx.get('out_gross') or 0
            all_time_ret = a_tx.get('ret_good') or 0
            all_time_ret_dmg = a_tx.get('ret_dmg') or 0
            all_time_dmg_std = a_tx.get('dmg_std') or 0
            all_time_adj = a_tx.get('adj') or 0
            
            # Net sold all-time = gross sold all-time - good returns all-time - damaged returns all-time
            all_time_out_net = all_time_out_gross - all_time_ret - all_time_ret_dmg
            # Cumulative received till now = current stock + all-time gross out + all-time standard damaged - all-time good returns - all-time adj
            total_rec = curr_stock + all_time_out_gross + all_time_dmg_std - all_time_ret - all_time_adj
            
            # Calculate closing stock at end_date if in the past
            if end_date < today:
                f_tx = future_map.get((p.id, b.id), {})
                future_in_std = f_tx.get('in_std') or 0
                future_ret = f_tx.get('ret_good') or 0
                future_out_gross = f_tx.get('out_gross') or 0
                future_dmg_std = f_tx.get('dmg_std') or 0
                future_adj = f_tx.get('adj') or 0
                
                closing_stock = curr_stock - (future_in_std + future_ret) + future_out_gross + future_dmg_std - future_adj
            else:
                closing_stock = curr_stock
                
            # Compute opening stock: Cl_stock - (In + Ret) + Out_gross + Dmg_std - Adj
            opening_stock = closing_stock - (period_in_std + period_ret) + period_out_gross + period_dmg_std - period_adj
            
            # Net sold in period = gross out - good returns - damaged returns
            period_out_net = period_out_gross - period_ret - period_ret_dmg
            
            if opening_stock == 0 and period_in_std == 0 and period_ret == 0 and period_out_net == 0 and period_dmg_std == 0 and period_ret_dmg == 0 and period_adj == 0 and closing_stock == 0 and (p.id, b.id) not in stock_map:
                continue
            
            # Net all-time damaged = standard damaged all-time + return damaged all-time
            all_time_dmg_net = all_time_dmg_std + all_time_ret_dmg

            item['total_op'] += opening_stock
            item['total_in'] += period_in_std
            item['total_ret'] += period_ret
            item['total_out'] += period_out_net
            item['total_adj_plus'] += period_adj_plus
            item['total_adj_minus'] += period_adj_minus
            item['total_dmg'] += period_dmg_std
            item['total_ret_dmg'] += period_ret_dmg
            item['total_cl'] += closing_stock
            item['total_rec'] += total_rec
            item['total_all_time_out'] += all_time_out_net
            item['total_all_time_dmg'] += all_time_dmg_net
            
            item['branch_stocks'].append({
                'branch': b,
                'op': opening_stock,
                'in': period_in_std,
                'ret': period_ret,
                'out': period_out_net,
                'adj_plus': period_adj_plus,
                'adj_minus': period_adj_minus,
                'dmg': period_dmg_std,
                'ret_dmg': period_ret_dmg,
                'cl': closing_stock,
                'rec': total_rec,
                'all_time_out': all_time_out_net,
                'all_time_dmg': all_time_dmg_net,
                'reg_id': reg_id_map.get((p.id, b.id)),
            })
            
        report_data.append(item)
        
    return render(request, 'core/stock_pivot_report.html', {
        'branches': active_branches,
        'accessible_branches': accessible_branches,
        'selected_branch': selected_branch,
        'report_data': report_data,
        'page_obj': page_obj,
        'start_date': start_date.isoformat(),
        'end_date': end_date.isoformat(),
        'show_op': True,
        'show_in': True,
        'show_out': True,
    })

@login_required
def export_stock_pivot_excel(request):
    """Export the multi-branch stock pivot report to an Excel file with multiple sheets."""
    if request.user.role == 'sales_staff':
        return redirect('dashboard')
        
    # Mirroring logic from view
    from django.db.models import Sum, Q
    import datetime
    import openpyxl
    from django.http import HttpResponse
    from django.utils import timezone

    q = request.GET.get('q', '').strip()
    from_date_str = request.GET.get('from_date', '')
    to_date_str = request.GET.get('to_date', '')
    
    today = timezone.now().date()
    start_date = today
    end_date = today
    
    if from_date_str:
        try:
            start_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
            
    if to_date_str:
        try:
            end_date = datetime.datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    accessible_branches = request.user.get_accessible_branches().order_by('name')
    branch_id = request.GET.get('branch', '').strip()
    
    selected_branch = None
    if branch_id:
        try:
            selected_branch = accessible_branches.get(id=int(branch_id))
        except (ValueError, Branch.DoesNotExist):
            pass

    if selected_branch:
        active_branches = [selected_branch]
    else:
        active_branches = list(accessible_branches)

    if q:
        products = Product.objects.filter(
            Q(barcode__icontains=q) | Q(name__icontains=q),
            branches__in=active_branches
        ).distinct().order_by('name')
    else:
        products = Product.objects.filter(branches__in=active_branches).distinct().order_by('name')

    registries = ProductRegistry.objects.filter(branch__in=active_branches).values('product_id', 'branch_id', 'stock_quantity')
    stock_map = {(r['product_id'], r['branch_id']): r['stock_quantity'] for r in registries}

    start_datetime = timezone.make_aware(datetime.datetime.combine(start_date, datetime.time.min))
    end_datetime = timezone.make_aware(datetime.datetime.combine(end_date, datetime.time.max))

    # Period transactions using conditional aggregation
    period_txns = StockTransaction.objects.filter(
        created_at__range=(start_datetime, end_datetime),
        branch__in=active_branches
    ).values('product_id', 'branch_id').annotate(
        in_std=Sum('quantity', filter=Q(transaction_type='IN') & ~Q(reference__startswith='Return #')),
        ret_good=Sum('quantity', filter=Q(transaction_type='IN') & Q(reference__startswith='Return #')),
        out_gross=Sum('quantity', filter=Q(transaction_type='OUT')),
        dmg_std=Sum('quantity', filter=Q(transaction_type='DMG') & ~Q(reference__startswith='Return #')),
        ret_dmg=Sum('quantity', filter=Q(transaction_type='DMG') & Q(reference__startswith='Return #')),
        adj=Sum('quantity', filter=Q(transaction_type='ADJ')),
    )
    period_map = {(t['product_id'], t['branch_id']): t for t in period_txns}

    # All-time transactions using conditional aggregation
    all_time_txns = StockTransaction.objects.filter(
        branch__in=active_branches
    ).values('product_id', 'branch_id').annotate(
        out_gross=Sum('quantity', filter=Q(transaction_type='OUT')),
        ret_good=Sum('quantity', filter=Q(transaction_type='IN') & Q(reference__startswith='Return #')),
        ret_dmg=Sum('quantity', filter=Q(transaction_type='DMG') & Q(reference__startswith='Return #')),
        dmg_std=Sum('quantity', filter=Q(transaction_type='DMG') & ~Q(reference__startswith='Return #')),
        adj=Sum('quantity', filter=Q(transaction_type='ADJ')),
    )
    all_time_map = {(t['product_id'], t['branch_id']): t for t in all_time_txns}
        
    future_map = {}
    if end_date < today:
        future_txns = StockTransaction.objects.filter(
            created_at__gt=end_datetime,
            branch__in=active_branches
        ).values('product_id', 'branch_id').annotate(
            in_std=Sum('quantity', filter=Q(transaction_type='IN') & ~Q(reference__startswith='Return #')),
            ret_good=Sum('quantity', filter=Q(transaction_type='IN') & Q(reference__startswith='Return #')),
            out_gross=Sum('quantity', filter=Q(transaction_type='OUT')),
            dmg_std=Sum('quantity', filter=Q(transaction_type='DMG') & ~Q(reference__startswith='Return #')),
            adj=Sum('quantity', filter=Q(transaction_type='ADJ')),
        )
        future_map = {(t['product_id'], t['branch_id']): t for t in future_txns}

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="stock_report_{start_date}_to_{end_date}.xlsx"'

    wb = openpyxl.Workbook()
    
    # Sheet 1: Consolidated Totals
    ws1 = wb.active
    ws1.title = "Consolidated Totals"
    header1 = [
        'Product Name', 'Barcode', 'Price', 
        'Total Op. Qty', 'Total In Qty', 'Total Ret. Qty', 'Total Out Qty', 
        'Total + Adj Qty', 'Total - Adj Qty', 'Total Damaged Qty', 'Total Ret. Damaged Qty', 
        'Total Cl. Qty', 'Total Received Qty', 'Total Sold (All-Time)', 'Total Damaged (All-Time)'
    ]
    ws1.append(header1)
    
    # Sheet 2: Individual Branch Data
    ws2 = wb.create_sheet(title="Branch Details")
    header2 = [
        'Product', 'Barcode', 'Branch Name', 'Branch Code', 
        'Op Qty', 'In Qty', 'Ret Qty', 'Out Qty', 
        '+ Adj Qty', '- Adj Qty', 'Damaged Qty', 'Ret Damaged Qty', 
        'Cl Qty', 'Total Received Qty', 'Total Sold (All-Time)', 'Total Damaged (All-Time)'
    ]
    ws2.append(header2)

    for p in products:
        total_op = 0
        total_in = 0
        total_ret = 0
        total_out = 0
        total_adj_plus = 0
        total_adj_minus = 0
        total_dmg = 0
        total_ret_dmg = 0
        total_cl = 0
        total_rec_sum = 0
        total_all_time_out_sum = 0
        total_all_time_dmg_sum = 0
        
        for b in active_branches:
            curr_stock = stock_map.get((p.id, b.id), 0)
            
            p_tx = period_map.get((p.id, b.id), {})
            period_in_std = p_tx.get('in_std') or 0
            period_ret = p_tx.get('ret_good') or 0
            period_out_gross = p_tx.get('out_gross') or 0
            period_dmg_std = p_tx.get('dmg_std') or 0
            period_ret_dmg = p_tx.get('ret_dmg') or 0
            period_adj = p_tx.get('adj') or 0
            
            period_adj_plus = period_adj if period_adj > 0 else 0
            period_adj_minus = abs(period_adj) if period_adj < 0 else 0
            
            # All-time aggregates for all-time columns
            a_tx = all_time_map.get((p.id, b.id), {})
            all_time_out_gross = a_tx.get('out_gross') or 0
            all_time_ret = a_tx.get('ret_good') or 0
            all_time_ret_dmg = a_tx.get('ret_dmg') or 0
            all_time_dmg_std = a_tx.get('dmg_std') or 0
            all_time_adj = a_tx.get('adj') or 0
            
            # Net sold all-time = gross sold all-time - good returns all-time - damaged returns all-time
            all_time_out_net = all_time_out_gross - all_time_ret - all_time_ret_dmg
            # Cumulative received till now = current stock + all-time gross out + all-time standard damaged - all-time good returns - all-time adj
            total_rec = curr_stock + all_time_out_gross + all_time_dmg_std - all_time_ret - all_time_adj
            
            # Calculate closing stock at end_date if in the past
            if end_date < today:
                f_tx = future_map.get((p.id, b.id), {})
                future_in_std = f_tx.get('in_std') or 0
                future_ret = f_tx.get('ret_good') or 0
                future_out_gross = f_tx.get('out_gross') or 0
                future_dmg_std = f_tx.get('dmg_std') or 0
                future_adj = f_tx.get('adj') or 0
                
                closing_stock = curr_stock - (future_in_std + future_ret) + future_out_gross + future_dmg_std - future_adj
            else:
                closing_stock = curr_stock
                
            # Compute opening stock: Cl_stock - (In + Ret) + Out_gross + Dmg_std - Adj
            opening_stock = closing_stock - (period_in_std + period_ret) + period_out_gross + period_dmg_std - period_adj
            
            # Net sold in period = gross out - good returns - damaged returns
            period_out_net = period_out_gross - period_ret - period_ret_dmg
            
            if opening_stock == 0 and period_in_std == 0 and period_ret == 0 and period_out_net == 0 and period_dmg_std == 0 and period_ret_dmg == 0 and period_adj == 0 and closing_stock == 0 and (p.id, b.id) not in stock_map:
                continue
            
            # Net all-time damaged = standard damaged + return damaged
            all_time_dmg_net = all_time_dmg_std + all_time_ret_dmg

            total_op += opening_stock
            total_in += period_in_std
            total_ret += period_ret
            total_out += period_out_net
            total_adj_plus += period_adj_plus
            total_adj_minus += period_adj_minus
            total_dmg += period_dmg_std
            total_ret_dmg += period_ret_dmg
            total_cl += closing_stock
            total_rec_sum += total_rec
            total_all_time_out_sum += all_time_out_net
            total_all_time_dmg_sum += all_time_dmg_net
            
            row2 = [
                p.name, p.barcode, b.name, b.code or '', 
                opening_stock, period_in_std, period_ret, period_out_net, 
                period_adj_plus, period_adj_minus, period_dmg_std, period_ret_dmg, 
                closing_stock, total_rec, all_time_out_net, all_time_dmg_net
            ]
            ws2.append(row2)
            
        row1 = [
            p.name, p.barcode, str(int(p.price)), 
            total_op, total_in, total_ret, total_out, 
            total_adj_plus, total_adj_minus, total_dmg, total_ret_dmg, 
            total_cl, total_rec_sum, total_all_time_out_sum, total_all_time_dmg_sum
        ]
        ws1.append(row1)

    # Sheet 3: Detailed Ledger
    ws3 = wb.create_sheet(title="Detailed Ledger")
    header3 = ['Date', 'Product', 'Barcode', 'Branch Name', 'Branch Code', 'Type', 'Quantity', 'Reference', 'User']
    ws3.append(header3)

    raw_txns = StockTransaction.objects.filter(
        created_at__gte=start_datetime,
        created_at__lte=end_datetime,
        branch__in=active_branches
    ).select_related('product', 'branch', 'user').order_by('created_at')

    for t in raw_txns:
        user_name = t.user.username if t.user else 'System'
        local_date = timezone.localtime(t.created_at).strftime("%Y-%m-%d %H:%M:%S")
        ws3.append([
            local_date,
            t.product.name,
            t.product.barcode,
            t.branch.name,
            t.branch.code or '',
            t.get_transaction_type_display(),
            t.quantity,
            t.reference or '',
            user_name
        ])

    wb.save(response)
    return response


@login_required
def stock_adjustment(request, reg_id):
    """
    Wrong-entry correction using reconciliation:
        Closing = Opening Balance + Stock In - Stock Out + Correction
    """
    from django.db.models import Sum, Q
    import datetime
    from django.utils import timezone
    from .models import StockAdjustment

    if not (request.user.role == 'owner' or request.user.has_product_rights):
        messages.error(request, "Permission denied. You do not have product edit rights.")
        return redirect('product_list')


    registry = get_object_or_404(ProductRegistry, pk=reg_id)
    product = registry.product
    branch = registry.branch

    today = timezone.now().date()
    start_datetime = timezone.make_aware(datetime.datetime.combine(today, datetime.time.min))

    # Aggregate today's IN, OUT, ADJ transactions for this product/branch
    txns = StockTransaction.objects.filter(
        product=product,
        branch=branch,
        created_at__gte=start_datetime,
    ).values('transaction_type').annotate(total=Sum('quantity'))

    period_in = 0
    period_out = 0
    for t in txns:
        if t['transaction_type'] in ('IN', 'ADJ'):
            period_in += t['total']
        elif t['transaction_type'] == 'OUT':
            period_out += t['total']

    current_stock = registry.stock_quantity  # This is the running closing stock
    opening_balance = current_stock - period_in + period_out

    error_msg = None

    if request.method == 'POST':
        try:
            input_amount = int(request.POST.get('correction_amount', 0))
            is_damage = request.POST.get('is_damage') == 'true'
            reason = request.POST.get('reason', '').strip()
            
            if is_damage:
                correction_amount = -abs(input_amount)
                default_reason = 'Damaged Stock'
                txn_type = 'DMG'
                txn_ref = f"DAMAGE ({correction_amount}): {reason or default_reason}"
            else:
                correction_amount = input_amount
                default_reason = 'Wrong Entry Correction'
                txn_type = 'ADJ'
                txn_ref = f"CORRECTION ({'+' if correction_amount >= 0 else ''}{correction_amount}): {reason or default_reason}"
                
            final_reason = reason or default_reason

            # Compute new closing
            new_closing = opening_balance + period_in - period_out + correction_amount

            if new_closing < 0:
                error_msg = f"Correction would result in negative stock ({new_closing}). Please enter a valid correction."
            else:
                # Update registry
                registry.stock_quantity = new_closing
                registry.save()

                is_in_stock = new_closing > 0

                # Audit record
                StockAdjustment.objects.create(
                    product=product,
                    branch=branch,
                    opening_balance=opening_balance,
                    stock_in=period_in,
                    stock_out=period_out,
                    correction_amount=correction_amount,
                    closing_stock=new_closing,
                    is_in_stock=is_in_stock,
                    reason=final_reason,
                    user=request.user,
                )

                # Ledger entry (signed quantity for ADJ/DMG)
                if correction_amount != 0:
                    StockTransaction.objects.create(
                        product=product,
                        branch=branch,
                        transaction_type=txn_type,
                        quantity=correction_amount if txn_type == 'ADJ' else abs(correction_amount),
                        reference=txn_ref,
                        user=request.user,
                    )

                success_msg = f"Stock corrected for {product.name} at {branch.name}: new closing = {new_closing}"
                if is_damage:
                    success_msg = f"Logged damage and updated stock for {product.name} at {branch.name}: new closing = {new_closing}"
                messages.success(request, success_msg)
                return redirect('stock_pivot_report')

        except (ValueError, TypeError):
            error_msg = "Invalid correction amount. Please enter a whole number."

    # Compute preview closing (default: no correction)
    preview_closing = opening_balance + period_in - period_out  # == current_stock

    context = {
        'registry': registry,
        'product': product,
        'branch': branch,
        'opening_balance': opening_balance,
        'period_in': period_in,
        'period_out': period_out,
        'current_stock': current_stock,
        'preview_closing': preview_closing,
        'error_msg': error_msg,
        # recent adjustments for this product/branch
        'recent_adjustments': StockAdjustment.objects.filter(
            product=product, branch=branch
        )[:10],
    }
    return render(request, 'core/stock_adjustment.html', context)

@login_required
def view_stock_adjustments(request, reg_id):
    """
    Read-only view showing the history of all stock adjustments/corrections
    for a given product registry (product + branch).
    """
    if request.user.role == 'sales_staff':
        return redirect('dashboard')
        
    registry = get_object_or_404(ProductRegistry, pk=reg_id)
    
    # Branch-level authorization check
    if not request.user.get_accessible_branches().filter(id=registry.branch.id).exists():
        messages.error(request, "Permission denied. You do not have access to this branch.")
        return redirect('stock_pivot_report')

    adjustments = StockAdjustment.objects.filter(
        product=registry.product,
        branch=registry.branch
    ).select_related('user').order_by('-created_at')

    context = {
        'registry': registry,
        'product': registry.product,
        'branch': registry.branch,
        'adjustments': adjustments,
    }
    return render(request, 'core/view_stock_adjustments.html', context)

@login_required
def pos_view(request):
    # Similar to product_list but uses the POS template
    from django.db.models import Q, Count, Sum, F, DecimalField, ExpressionWrapper
    from django.core.paginator import Paginator

    q = request.GET.get('q', '')
    selected_branch = request.GET.get('branch', '')
    active_filter = request.GET.get('filter', '')

    # Determine accessible branches; if none, show all branches
    accessible_branches_qs = request.user.get_accessible_branches()
    if not accessible_branches_qs.exists():
        from core.models import Branch
        accessible_branches = Branch.objects.all()
    else:
        accessible_branches = accessible_branches_qs
    # Initialize registrations queryset
    registrations = ProductRegistry.objects.select_related('product', 'branch').filter(branch__in=accessible_branches)

    if q:
        registrations = registrations.filter(Q(product__name__icontains=q) | Q(product__barcode__icontains=q))

    if selected_branch:
        registrations = registrations.filter(branch_id=selected_branch)

    if active_filter == 'low_stock':
        registrations = registrations.filter(stock_quantity__lte=F('low_stock_threshold'))
    elif active_filter == 'zero_stock':
        registrations = registrations.filter(stock_quantity=0)

    registrations = registrations.order_by('product__name')

    # Basic stats (optional for POS)
    stats = registrations.aggregate(
        product_count=Count('product', distinct=True),
        registration_count=Count('id'),
        total_value=Sum(ExpressionWrapper(F('stock_quantity') * F('product__price'), output_field=DecimalField()))
    )

    paginator = Paginator(registrations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    branches = accessible_branches

    return render(request, 'pos/index.html', {
        'page_obj': page_obj,
        'registrations': page_obj.object_list,  # iterable for loop
        'products': registrations,  # original queryset for count
        'stats': stats,
        'q': q,
        'branches': branches,
        'selected_branch': selected_branch,
        'active_filter': active_filter,
    })


@login_required
def api_recent_bills(request):
    """Return recent bills as JSON for auto-refresh on dashboard."""
    # Optionally limit to recent 5 bills
    recent = Bill.objects.order_by('-created_at')[:5]
    bills = []
    for b in recent:
        bills.append({
            'id': b.id,
            'invoice_number': b.invoice_number,
            'branch_name': b.branch.name if b.branch else '',
            'total_amount': float(b.total_amount),
        })
    return JsonResponse({'bills': bills})

