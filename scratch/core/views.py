from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Product, Branch, ProductRegistry, StockTransaction
from django import forms
from .forms import ProductForm


@login_required
def product_list(request):
    if request.user.role == 'sales_staff':
        return redirect('dashboard')
    
    # Auto-initialize active branch if None
    if not request.user.active_branch:
        accessible = request.user.get_accessible_branches()
        if accessible.exists():
            request.user.active_branch = accessible.first()
            request.user.save()
    
    from django.db.models import Q, Count, Sum, F, DecimalField, ExpressionWrapper
    from django.core.paginator import Paginator

    q = request.GET.get('q', '')
    selected_branch = request.GET.get('branch', '')
    active_filter = request.GET.get('filter', '')
    
    # Base query
    accessible_branches = request.user.get_accessible_branches()
    registrations = ProductRegistry.objects.select_related('product', 'branch').prefetch_related('product__combos').filter(branch__in=accessible_branches)

    if q:
        registrations = registrations.filter(
            Q(product__name__icontains=q) | Q(product__barcode__icontains=q)
        )

    if selected_branch:
        registrations = registrations.filter(branch_id=selected_branch)
    
    if active_filter == 'low_stock':
        registrations = registrations.filter(stock_quantity__lte=F('low_stock_threshold'))
    elif active_filter == 'zero_stock':
        registrations = registrations.filter(stock_quantity=0)
        
    registrations = registrations.order_by('product__name')

    # Stats
    stats = registrations.aggregate(
        product_count=Count('product', distinct=True),
        registration_count=Count('id'),
        total_value=Sum(
            ExpressionWrapper(F('stock_quantity') * F('product__price'), output_field=DecimalField())
        )
    )

    # Pagination
    paginator = Paginator(registrations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    branches = accessible_branches

    return render(request, 'core/product_list.html', {
        'page_obj': page_obj,
        'stats': stats,
        'q': q,
        'branches': branches,
        'selected_branch': selected_branch,
        'active_filter': active_filter,
    })

@login_required
def export_products_csv(request):
    if request.user.role == 'sales_staff':
        return redirect('dashboard')
        
    import csv
    from django.http import HttpResponse

    q = request.GET.get('q', '')
    selected_branch = request.GET.get('branch', '')
    
    accessible_branches = request.user.get_accessible_branches()
    registrations = ProductRegistry.objects.select_related('product', 'branch').prefetch_related('product__combos').filter(branch__in=accessible_branches)

    if q:
        from django.db.models import Q
        registrations = registrations.filter(
            Q(product__name__icontains=q) | Q(product__barcode__icontains=q)
        )
    
    if selected_branch:
        registrations = registrations.filter(branch_id=selected_branch)

    registrations = registrations.order_by('-created_at')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'

    writer = csv.writer(response)
    writer.writerow(['Branch Name', 'Branch Code', 'Product Name', 'Barcode', 'Price', 'Combo Prices', 'Stock', 'Low Stock Level', 'Registered On'])

    for reg in registrations:
        date_str = reg.created_at.strftime("%Y-%m-%d %H:%M")
        combos = reg.product.combos.all()
        combo_str = ' | '.join([f"Qty:{int(c.quantity)} @ ₹{int(c.price)}" for c in combos]) if combos else '-'
        writer.writerow([
            reg.branch.name, 
            reg.branch.code or '',
            reg.product.name, 
            reg.product.barcode, 
            str(int(reg.product.price)),
            combo_str,
            reg.stock_quantity,
            reg.low_stock_threshold,
            date_str
        ])

    return response


@login_required
def product_create(request):
    if not (request.user.role == 'owner' or request.user.has_product_rights):
        from django.contrib import messages
        messages.error(request, "Permission denied. You do not have product edit rights.")
        return redirect('product_list')

    
    # Auto-initialize active branch if None
    if not request.user.active_branch:
        accessible = request.user.get_accessible_branches()
        if accessible.exists():
            request.user.active_branch = accessible.first()
            request.user.save()

    if request.method == 'POST':
        form = ProductForm(request.POST)
        from .forms import ComboPriceFormSet
        combo_formset = ComboPriceFormSet(request.POST)
        
        if form.is_valid() and combo_formset.is_valid():
            product = form.save()
            combo_formset.instance = product
            combo_formset.save()
            
            initial_branch = form.cleaned_data.get('initial_branch')
            initial_stock = form.cleaned_data.get('initial_stock') or 0
            low_threshold = form.cleaned_data.get('low_stock_threshold') or 10
            
            # Determine branch for registration (fallback to active_branch)
            branch = initial_branch or request.user.active_branch
            if branch:
                ProductRegistry.objects.create(
                    branch=branch,
                    product=product,
                    stock_quantity=initial_stock,
                    low_stock_threshold=low_threshold,
                )
                # Manual stock input maps to opening stock; no StockTransaction is recorded for this initial quantity.
                return redirect('product_list')
        else:
            # Check if form is invalid specifically because the product already exists
            barcode = request.POST.get('barcode', '').strip()
            if barcode:
                existing_product = Product.objects.filter(barcode=barcode).first()
                if existing_product:
                    initial_branch_id = request.POST.get('initial_branch')
                    initial_stock = int(request.POST.get('initial_stock') or 0)
                    low_threshold = int(request.POST.get('low_stock_threshold') or 10)
                    
                    if initial_branch_id:
                        branch = get_object_or_404(Branch, id=initial_branch_id)
                        reg, created = ProductRegistry.objects.get_or_create(
                            branch=branch,
                            product=existing_product,
                            defaults={
                                'stock_quantity': initial_stock,
                                'low_stock_threshold': low_threshold
                            }
                        )
                        
                        from django.contrib import messages
                        if created:
                            if initial_stock > 0:
                                StockTransaction.objects.create(
                                    product=existing_product,
                                    branch=branch,
                                    transaction_type='IN',
                                    quantity=initial_stock,
                                    reference='Initial Stock',
                                    user=request.user
                                )
                            messages.success(request, f'Product "{existing_product.name}" already exists and has been successfully registered to {branch.name}!')
                        else:
                            # Update stock if it was 0 or just warn
                            if reg.stock_quantity == 0 and initial_stock > 0:
                                reg.stock_quantity = initial_stock
                                reg.save()
                                StockTransaction.objects.create(
                                    product=existing_product,
                                    branch=branch,
                                    transaction_type='IN',
                                    quantity=initial_stock,
                                    reference='Stock Update',
                                    user=request.user
                                )
                                messages.success(request, f'Stock updated for "{existing_product.name}" at {branch.name}!')
                            else:
                                messages.warning(request, f'Product "{existing_product.name}" is already registered to {branch.name}.')
                        
                        return redirect('product_list')
    else:
        form = ProductForm(initial={
            'initial_branch': request.user.active_branch,
            'initial_stock': 0,
            'low_stock_threshold': 10
        })
        from .forms import ComboPriceFormSet
        combo_formset = ComboPriceFormSet()
    return render(request, 'core/product_form.html', {'form': form, 'combo_formset': combo_formset, 'action': 'Add New'})

@login_required
def product_update(request, pk):
    if not (request.user.role == 'owner' or request.user.has_product_rights):
        from django.contrib import messages
        messages.error(request, "Permission denied. You do not have product edit rights.")
        return redirect('product_list')

    product = get_object_or_404(Product, pk=pk)
    
    # For updating, we might want to edit a specific registration
    reg_id = request.GET.get('reg_id')
    registration = None
    if reg_id:
        registration = get_object_or_404(ProductRegistry, id=reg_id, product=product)

    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        from .forms import ComboPriceFormSet
        combo_formset = ComboPriceFormSet(request.POST, instance=product)
        
        if form.is_valid() and combo_formset.is_valid():
            product = form.save()
            combo_formset.save()
            
            # If we are editing stock for a specific registration
            if registration:
                old_stock = registration.stock_quantity
                new_stock = int(request.POST.get('stock_quantity', registration.stock_quantity))
                registration.stock_quantity = new_stock
                registration.low_stock_threshold = request.POST.get('low_stock_threshold', registration.low_stock_threshold)
                registration.save()
            
            return redirect('product_list')
    else:
        initial_data = {}
        if registration:
            initial_data = {
                'initial_branch': registration.branch,
                'initial_stock': registration.stock_quantity,
                'low_stock_threshold': registration.low_stock_threshold
            }
        form = ProductForm(instance=product, initial=initial_data)
        from .forms import ComboPriceFormSet
        combo_formset = ComboPriceFormSet(instance=product)
        
    return render(request, 'core/product_form.html', {
        'form': form, 
        'combo_formset': combo_formset,
        'action': 'Edit', 
        'registration': registration
    })

@login_required
def update_product_stock_ajax(request, pk):
    # This pk is the ProductRegistry ID
    if not (request.user.role == 'owner' or request.user.has_product_rights):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_stock = data.get('stock')
            if new_stock is not None:
                reg = get_object_or_404(ProductRegistry, pk=pk)
                old_stock = reg.stock_quantity
                reg.stock_quantity = int(new_stock)
                reg.save()
                
                if reg.stock_quantity != old_stock:
                    diff = reg.stock_quantity - old_stock
                    StockTransaction.objects.create(
                        product=reg.product,
                        branch=reg.branch,
                        transaction_type='IN' if diff > 0 else 'OUT',
                        quantity=abs(diff),
                        reference='Ajax Update',
                        user=request.user
                    )
                    
                return JsonResponse({'success': True, 'stock': reg.stock_quantity})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
            
    return JsonResponse({'error': 'Invalid request'}, status=405)

from django.http import JsonResponse
import json

@login_required
def update_product_price_ajax(request, pk):
    if not (request.user.role == 'owner' or request.user.has_product_rights):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_price = data.get('price')
            if new_price is not None:
                product = get_object_or_404(Product, pk=pk)
                product.price = new_price
                product.save()
                return JsonResponse({'success': True, 'price': float(product.price)})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
            
    return JsonResponse({'error': 'Invalid request'}, status=405)

import csv
import io
from django.contrib import messages

@login_required
def bulk_insert(request):
    if not (request.user.role == 'owner' or request.user.has_product_rights):
        from django.contrib import messages
        messages.error(request, "Permission denied. You do not have product edit rights.")
        return redirect('product_list')

    
    branches = request.user.get_accessible_branches()
    results = None
    
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            messages.error(request, 'Please upload a CSV file.')
            return render(request, 'core/bulk_insert.html', {'branches': branches})
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'File must be a .csv file.')
            return render(request, 'core/bulk_insert.html', {'branches': branches})
        
        try:
            decoded_file = csv_file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded_file))
            
            success_count = 0
            error_count = 0
            errors = []
            
            for row_num, row in enumerate(reader, start=2):
                try:
                    # Support both template formats:
                    # New template: Name, Barcode, Price, Size, Branch Code, Initial Stock, Low Stock Alert
                    # Old template: Branch, Product Name, Barcode, Size, Price, Stock, Low Stock Level
                    product_name = (row.get('Name') or row.get('Product Name') or '').strip()
                    barcode = (row.get('Barcode') or '').strip()
                    size = (row.get('Size') or '').strip()
                    price = (row.get('Price') or '0').strip()
                    stock = (row.get('Initial Stock') or row.get('Stock') or '0').strip()
                    low_stock = (row.get('Low Stock Alert') or row.get('Low Stock Level') or '10').strip()
                    branch_code = (row.get('Branch Code') or '').strip()
                    branch_name = (row.get('Branch') or '').strip()
                    
                    if not product_name or not barcode:
                        errors.append(f"Row {row_num}: Missing product name or barcode.")
                        error_count += 1
                        continue
                    
                    # Find or create branch (by code first, then by name)
                    branch = None
                    if branch_code:
                        try:
                            branch = Branch.objects.get(code=int(branch_code))
                        except (Branch.DoesNotExist, ValueError):
                            errors.append(f"Row {row_num}: Branch with code '{branch_code}' not found.")
                            error_count += 1
                            continue
                    elif branch_name:
                        branch, _ = Branch.objects.get_or_create(name=branch_name)
                    
                    # Find or create product
                    product, created = Product.objects.get_or_create(
                        barcode=barcode,
                        defaults={
                            'name': product_name,
                            'price': float(price) if price else 0,
                            'size': size if size else ''
                        }
                    )
                    
                    if not created:
                        # Update existing product price/name/size
                        product.name = product_name
                        product.price = float(price) if price else product.price
                        product.size = size if size else product.size
                        product.save()
                    
                    # Create registry entry if branch is provided
                    if branch:
                        reg, reg_created = ProductRegistry.objects.get_or_create(
                            branch=branch,
                            product=product,
                            defaults={
                                'stock_quantity': int(stock) if stock else 0,
                                'low_stock_threshold': int(low_stock) if low_stock else 10
                            }
                        )
                        if not reg_created:
                            old_stock = reg.stock_quantity
                            reg.stock_quantity = old_stock + (int(stock) if stock else 0)
                            reg.low_stock_threshold = int(low_stock) if low_stock else reg.low_stock_threshold
                            reg.save()
                            
                            if reg.stock_quantity != old_stock:
                                diff = reg.stock_quantity - old_stock
                                StockTransaction.objects.create(
                                    product=product,
                                    branch=branch,
                                    transaction_type='IN' if diff > 0 else 'OUT',
                                    quantity=abs(diff),
                                    reference='Bulk Update',
                                    user=request.user
                                )
                        else:
                            if reg.stock_quantity > 0:
                                StockTransaction.objects.create(
                                    product=product,
                                    branch=branch,
                                    transaction_type='IN',
                                    quantity=reg.stock_quantity,
                                    reference='Bulk Insert',
                                    user=request.user
                                )
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
            
            results = {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            }
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
    
    return render(request, 'core/bulk_insert.html', {
        'branches': branches,
        'results': results,
    })

@login_required
def download_bulk_template(request):
    # Updated bulk insert template header to include Branch Code
    header = ["Product Name","Barcode","Price","Size","Branch Code","Initial Stock","Low Stock Alert"]
    # This endpoint will generate a CSV template for bulk insert
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bulk_insert_template.csv"'
    writer = csv.writer(response)
    writer.writerow(header)
    writer.writerow(['Example Product', 'EX001', '299.00', 'M', 'BR001', '50', '10'])
    writer.writerow(['Another Product', 'EX002', '149.00', 'L', 'BR001', '100', '15'])
    
    return response

@login_required
def stock_pivot_report(request):
    """Report showing products, their stock movement (Op, In, Out, Cl) and branch-wise closing stock."""
    if request.user.role == 'sales_staff':
        return redirect('dashboard')
    
    from django.db.models import Sum, Q, F
    import datetime
    from django.utils import timezone
    from django.core.paginator import Paginator
    
    q = request.GET.get('q', '').strip()
    from_date_str = request.GET.get('from_date', '')
    to_date_str = request.GET.get('to_date', '')
    
    today = timezone.now().date()
    start_date = today
    end_date = today
    
    if from_date_str and to_date_str:
        try:
            start_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
        
    branches = list(request.user.get_accessible_branches().order_by('name'))
    active_branches = branches

    if q:
        products = Product.objects.filter(
            Q(barcode__icontains=q) | Q(name__icontains=q),
            branches__in=branches
        ).distinct().order_by('name')
    else:
        products = Product.objects.filter(branches__in=branches).distinct().order_by('name')

    # Pagination: 10 products per page
    paginator = Paginator(products, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    print("DEBUG PAGINATION: total products:", paginator.count, "has_other:", page_obj.has_other_pages())

    if request.GET.get('test_error'):
        raise Exception(f"Pagination triggered: {paginator.count} items, returning {len(page_obj.object_list)}")

    # 1. Fetch current stock via registries (restricted to accessible branches)
    registries = ProductRegistry.objects.filter(branch__in=branches).values('id', 'product_id', 'branch_id', 'stock_quantity')
    stock_map = {(r['product_id'], r['branch_id']): r['stock_quantity'] for r in registries}
    reg_id_map = {(r['product_id'], r['branch_id']): r['id'] for r in registries}
    
    # 2. Fetch stock movements
    start_datetime = timezone.make_aware(datetime.datetime.combine(start_date, datetime.time.min))
    end_datetime = timezone.make_aware(datetime.datetime.combine(end_date, datetime.time.max))

    txns = StockTransaction.objects.filter(
        created_at__gte=start_datetime,
        branch__in=branches
    ).values('product_id', 'branch_id', 'transaction_type').annotate(total_qty=Sum('quantity'))
    
    txn_map = {}
    for t in txns:
        key = (t['product_id'], t['branch_id'], t['transaction_type'])
        txn_map[key] = t['total_qty']
        
    future_txns_map = {}
    if end_date < today:
        future_txns = StockTransaction.objects.filter(
            created_at__gt=end_datetime,
            branch__in=branches
        ).values('product_id', 'branch_id', 'transaction_type').annotate(total_qty=Sum('quantity'))
        for t in future_txns:
            key = (t['product_id'], t['branch_id'], t['transaction_type'])
            future_txns_map[key] = t['total_qty']

    report_data = []
    
    for p in page_obj:
        item = {
            'product': p,
            'total_op': 0,
            'total_in': 0,
            'total_out': 0,
            'total_dmg': 0,
            'total_cl': 0,
            'branch_stocks': []
        }
        
        for b in active_branches:
            curr_stock = stock_map.get((p.id, b.id), 0)
            
            period_in = txn_map.get((p.id, b.id, 'IN'), 0) + txn_map.get((p.id, b.id, 'ADJ'), 0)
            period_out = txn_map.get((p.id, b.id, 'OUT'), 0)
            period_dmg = txn_map.get((p.id, b.id, 'DMG'), 0)
            
            future_in = future_txns_map.get((p.id, b.id, 'IN'), 0) + future_txns_map.get((p.id, b.id, 'ADJ'), 0)
            future_out = future_txns_map.get((p.id, b.id, 'OUT'), 0)
            future_dmg = future_txns_map.get((p.id, b.id, 'DMG'), 0)
            
            # Compute opening stock based on current stock and period transactions
            # Correct formula: opening = current_stock - period_in + period_out + period_dmg
            opening_stock = curr_stock - period_in + period_out + period_dmg
            # Closing stock after applying period transactions
            closing_stock = opening_stock + period_in - period_out - period_dmg
            
            if opening_stock == 0 and period_in == 0 and period_out == 0 and period_dmg == 0 and closing_stock == 0 and (p.id, b.id) not in stock_map:
                continue
            
            item['total_op'] += opening_stock
            item['total_in'] += period_in
            item['total_out'] += period_out
            item['total_dmg'] += period_dmg
            item['total_cl'] += closing_stock
            
            item['branch_stocks'].append({
                'branch': b,
                'op': opening_stock,
                'in': period_in,
                'out': period_out,
                'dmg': period_dmg,
                'cl': closing_stock,
                'reg_id': reg_id_map.get((p.id, b.id)),
            })
            
        report_data.append(item)
        
    return render(request, 'core/stock_pivot_report.html', {
        'branches': active_branches,
        'report_data': report_data,
        'page_obj': page_obj,
        'start_date': start_date.isoformat(),
        'end_date': end_date.isoformat(),
        'show_op': True,
        'show_in': True,
        'show_out': True,
    })

@login_required
def export_stock_pivot_excel(request):
    """Export the multi-branch stock pivot report to an Excel file with multiple sheets."""
    if request.user.role == 'staff':
        return redirect('dashboard')
        
    # Mirroring logic from view
    from django.db.models import Sum, Q
    import datetime
    import openpyxl
    from django.http import HttpResponse
    from django.utils import timezone

    q = request.GET.get('q', '').strip()
    from_date_str = request.GET.get('from_date', '')
    to_date_str = request.GET.get('to_date', '')
    
    today = timezone.now().date()
    start_date = today
    end_date = today
    
    if from_date_str and to_date_str:
        try:
            start_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    branches = list(request.user.get_accessible_branches().order_by('name'))
    active_branches = branches

    if q:
        products = Product.objects.filter(
            Q(barcode__icontains=q) | Q(name__icontains=q),
            branches__in=branches
        ).distinct().order_by('name')
    else:
        products = Product.objects.filter(branches__in=branches).distinct().order_by('name')

    registries = ProductRegistry.objects.filter(branch__in=branches).values('product_id', 'branch_id', 'stock_quantity')
    stock_map = {(r['product_id'], r['branch_id']): r['stock_quantity'] for r in registries}

    start_datetime = timezone.make_aware(datetime.datetime.combine(start_date, datetime.time.min))
    end_datetime = timezone.make_aware(datetime.datetime.combine(end_date, datetime.time.max))

    txns = StockTransaction.objects.filter(created_at__gte=start_datetime, branch__in=branches).values('product_id', 'branch_id', 'transaction_type').annotate(total_qty=Sum('quantity'))
    txn_map = {(t['product_id'], t['branch_id'], t['transaction_type']): t['total_qty'] for t in txns}
    
    future_txns_map = {}
    if end_date < today:
        future_txns = StockTransaction.objects.filter(created_at__gt=end_datetime, branch__in=branches).values('product_id', 'branch_id', 'transaction_type').annotate(total_qty=Sum('quantity'))
        future_txns_map = {(t['product_id'], t['branch_id'], t['transaction_type']): t['total_qty'] for t in future_txns}

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="stock_report_{start_date}_to_{end_date}.xlsx"'

    wb = openpyxl.Workbook()
    
    # Sheet 1: Consolidated Totals
    ws1 = wb.active
    ws1.title = "Consolidated Totals"
    header1 = ['Product Name', 'Barcode', 'Price', 'Total Op. Qty', 'Total In Qty', 'Total Out Qty', 'Total Damaged Qty', 'Total Cl. Qty']
    ws1.append(header1)
    
    # Sheet 2: Individual Branch Data
    ws2 = wb.create_sheet(title="Branch Details")
    header2 = ['Product', 'Barcode', 'Branch Name', 'Branch Code', 'Op Qty', 'In Qty', 'Out Qty', 'Damaged Qty', 'Cl Qty']
    ws2.append(header2)

    for p in products:
        total_op = 0
        total_in = 0
        total_out = 0
        total_dmg = 0
        total_cl = 0
        
        for b in active_branches:
            curr_stock = stock_map.get((p.id, b.id), 0)
            period_in = txn_map.get((p.id, b.id, 'IN'), 0) + txn_map.get((p.id, b.id, 'ADJ'), 0)
            period_out = txn_map.get((p.id, b.id, 'OUT'), 0)
            period_dmg = txn_map.get((p.id, b.id, 'DMG'), 0)
            
            future_in = future_txns_map.get((p.id, b.id, 'IN'), 0) + future_txns_map.get((p.id, b.id, 'ADJ'), 0)
            future_out = future_txns_map.get((p.id, b.id, 'OUT'), 0)
            future_dmg = future_txns_map.get((p.id, b.id, 'DMG'), 0)
            
            # Compute opening stock based on current stock and period transactions
            # Correct formula: opening = current_stock - period_in + period_out + period_dmg
            opening_stock = curr_stock - period_in + period_out + period_dmg
            # Closing stock after applying period transactions
            closing_stock = opening_stock + period_in - period_out - period_dmg
            
            if opening_stock == 0 and period_in == 0 and period_out == 0 and period_dmg == 0 and closing_stock == 0 and (p.id, b.id) not in stock_map:
                continue
            
            total_op += opening_stock
            total_in += period_in
            total_out += period_out
            total_dmg += period_dmg
            total_cl += closing_stock
            
            row2 = [p.name, p.barcode, b.name, b.code or '', opening_stock, period_in, period_out, period_dmg, closing_stock]
            ws2.append(row2)
            
        row1 = [p.name, p.barcode, str(int(p.price)), total_op, total_in, total_out, total_dmg, total_cl]
        ws1.append(row1)

    # Sheet 3: Detailed Ledger
    ws3 = wb.create_sheet(title="Detailed Ledger")
    header3 = ['Date', 'Product', 'Barcode', 'Branch Name', 'Branch Code', 'Type', 'Quantity', 'Reference', 'User']
    ws3.append(header3)

    raw_txns = StockTransaction.objects.filter(
        created_at__gte=start_datetime,
        created_at__lte=end_datetime,
        branch__in=branches
    ).select_related('product', 'branch', 'user').order_by('created_at')

    for t in raw_txns:
        user_name = t.user.username if t.user else 'System'
        local_date = timezone.localtime(t.created_at).strftime("%Y-%m-%d %H:%M:%S")
        ws3.append([
            local_date,
            t.product.name,
            t.product.barcode,
            t.branch.name,
            t.branch.code or '',
            t.get_transaction_type_display(),
            t.quantity,
            t.reference or '',
            user_name
        ])

    wb.save(response)
    return response


@login_required
def stock_adjustment(request, reg_id):
    """
    Wrong-entry correction using reconciliation:
        Closing = Opening Balance + Stock In - Stock Out + Correction
    """
    from django.db.models import Sum, Q
    import datetime
    from django.utils import timezone
    from .models import StockAdjustment

    if not (request.user.role == 'owner' or request.user.has_product_rights):
        from django.contrib import messages
        messages.error(request, "Permission denied. You do not have product edit rights.")
        return redirect('product_list')


    registry = get_object_or_404(ProductRegistry, pk=reg_id)
    product = registry.product
    branch = registry.branch

    today = timezone.now().date()
    start_datetime = timezone.make_aware(datetime.datetime.combine(today, datetime.time.min))

    # Aggregate today's IN, OUT, ADJ transactions for this product/branch
    txns = StockTransaction.objects.filter(
        product=product,
        branch=branch,
        created_at__gte=start_datetime,
    ).values('transaction_type').annotate(total=Sum('quantity'))

    period_in = 0
    period_out = 0
    for t in txns:
        if t['transaction_type'] in ('IN', 'ADJ'):
            period_in += t['total']
        elif t['transaction_type'] == 'OUT':
            period_out += t['total']

    current_stock = registry.stock_quantity  # This is the running closing stock
    opening_balance = current_stock - period_in + period_out

    error_msg = None

    if request.method == 'POST':
        try:
            input_amount = int(request.POST.get('correction_amount', 0))
            is_damage = request.POST.get('is_damage') == 'true'
            reason = request.POST.get('reason', '').strip()
            
            if is_damage:
                correction_amount = -abs(input_amount)
                default_reason = 'Damaged Stock'
                txn_type = 'DMG'
                txn_ref = f"DAMAGE ({correction_amount}): {reason or default_reason}"
            else:
                correction_amount = input_amount
                default_reason = 'Wrong Entry Correction'
                txn_type = 'ADJ'
                txn_ref = f"CORRECTION ({'+' if correction_amount >= 0 else ''}{correction_amount}): {reason or default_reason}"
                
            final_reason = reason or default_reason

            # Compute new closing
            new_closing = opening_balance + period_in - period_out + correction_amount

            if new_closing < 0:
                error_msg = f"Correction would result in negative stock ({new_closing}). Please enter a valid correction."
            else:
                # Update registry
                registry.stock_quantity = new_closing
                registry.save()

                is_in_stock = new_closing > 0

                # Audit record
                StockAdjustment.objects.create(
                    product=product,
                    branch=branch,
                    opening_balance=opening_balance,
                    stock_in=period_in,
                    stock_out=period_out,
                    correction_amount=correction_amount,
                    closing_stock=new_closing,
                    is_in_stock=is_in_stock,
                    reason=final_reason,
                    user=request.user,
                )

                # Ledger entry (signed quantity for ADJ/DMG)
                if correction_amount != 0:
                    StockTransaction.objects.create(
                        product=product,
                        branch=branch,
                        transaction_type=txn_type,
                        quantity=abs(correction_amount),
                        reference=txn_ref,
                        user=request.user,
                    )

                success_msg = f"Stock corrected for {product.name} at {branch.name}: new closing = {new_closing}"
                if is_damage:
                    success_msg = f"Logged damage and updated stock for {product.name} at {branch.name}: new closing = {new_closing}"
                messages.success(request, success_msg)
                return redirect('stock_pivot_report')

        except (ValueError, TypeError):
            error_msg = "Invalid correction amount. Please enter a whole number."

    # Compute preview closing (default: no correction)
    preview_closing = opening_balance + period_in - period_out  # == current_stock

    context = {
        'registry': registry,
        'product': product,
        'branch': branch,
        'opening_balance': opening_balance,
        'period_in': period_in,
        'period_out': period_out,
        'current_stock': current_stock,
        'preview_closing': preview_closing,
        'error_msg': error_msg,
        # recent adjustments for this product/branch
        'recent_adjustments': StockAdjustment.objects.filter(
            product=product, branch=branch
        )[:10],
    }
    return render(request, 'core/stock_adjustment.html', context)

@login_required
def pos_view(request):
    # Similar to product_list but uses the POS template
    from django.db.models import Q, Count, Sum, F, DecimalField, ExpressionWrapper
    from django.core.paginator import Paginator

    q = request.GET.get('q', '')
    selected_branch = request.GET.get('branch', '')
    active_filter = request.GET.get('filter', '')

    # Determine accessible branches; if none, show all branches
    accessible_branches_qs = request.user.get_accessible_branches()
    if not accessible_branches_qs.exists():
        from core.models import Branch
        accessible_branches = Branch.objects.all()
    else:
        accessible_branches = accessible_branches_qs
    # Initialize registrations queryset
    registrations = ProductRegistry.objects.select_related('product', 'branch').filter(branch__in=accessible_branches)

    if q:
        registrations = registrations.filter(Q(product__name__icontains=q) | Q(product__barcode__icontains=q))

    if selected_branch:
        registrations = registrations.filter(branch_id=selected_branch)

    if active_filter == 'low_stock':
        registrations = registrations.filter(stock_quantity__lte=F('low_stock_threshold'))
    elif active_filter == 'zero_stock':
        registrations = registrations.filter(stock_quantity=0)

    registrations = registrations.order_by('product__name')

    # Basic stats (optional for POS)
    stats = registrations.aggregate(
        product_count=Count('product', distinct=True),
        registration_count=Count('id'),
        total_value=Sum(ExpressionWrapper(F('stock_quantity') * F('product__price'), output_field=DecimalField()))
    )

    paginator = Paginator(registrations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    branches = accessible_branches

    return render(request, 'pos/index.html', {
        'page_obj': page_obj,
        'registrations': page_obj.object_list,  # iterable for loop
        'products': registrations,  # original queryset for count
        'stats': stats,
        'q': q,
        'branches': branches,
        'selected_branch': selected_branch,
        'active_filter': active_filter,
    })

